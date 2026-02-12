import json
import os
import re
import subprocess
import time
from copy import deepcopy
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import streamlit as st

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - runtime optional dependency
    load_workbook = None


def resolve_repo_root() -> Path:
    env_root = str(os.environ.get("SPEC_FACTORY_ROOT", "")).strip()
    if env_root:
        root = Path(env_root).resolve()
        if root.exists():
            return root
    return Path(__file__).resolve().parents[2]


REPO_ROOT = resolve_repo_root()
OUTPUT_ROOT = (REPO_ROOT / os.environ.get("LOCAL_OUTPUT_ROOT", "out")).resolve()
HELPER_ROOT = REPO_ROOT / "helper_files"
FINAL_ROOT = OUTPUT_ROOT / "final"
EVENTS_PATH = OUTPUT_ROOT / "_runtime" / "events.jsonl"
GUI_PROCESS_LOG_PATH = OUTPUT_ROOT / "_runtime" / "gui_process.log"
QUEUE_ROOT = OUTPUT_ROOT / "_queue"
RUNS_ROOT = OUTPUT_ROOT / "runs"
BILLING_ROOT = OUTPUT_ROOT / "_billing"
LEARNING_ROOT = OUTPUT_ROOT / "_learning"
COMPONENT_ROOT = OUTPUT_ROOT / "_components"
UNKNOWN = {"", "unk", "unknown", "na", "n/a", "none", "null"}

EVENT_MEANINGS = {
    "queue_transition": "Product changed queue state (pending/running/complete/exhausted).",
    "run_started": "Pipeline run started for the selected product.",
    "helper_files_context_loaded": "Helper files were loaded and matched against the product identity.",
    "helper_supportive_fill_applied": "Known supportive helper values were injected for missing fields.",
    "discovery_results_reranked": "Search/discovery results were reranked for best candidates.",
    "source_discovery_only": "Source was discovered but not fetched (planning/discovery step only).",
    "source_fetch_started": "A source URL fetch has started.",
    "source_processed": "A fetched source was parsed and candidate fields were extracted.",
    "source_fetch_failed": "A fetch failed (timeout, blocked, network, parser, etc).",
    "llm_call_started": "LLM call started (plan/extract/validate/summary).",
    "llm_call_usage": "LLM token/cost usage was recorded.",
    "llm_call_completed": "LLM call returned successfully.",
    "llm_call_failed": "LLM call failed for the configured provider.",
    "llm_discovery_planner_failed": "LLM discovery planner failed; deterministic fallback used.",
    "llm_extract_failed": "LLM extraction failed; deterministic extraction still continues.",
    "llm_extract_skipped_budget": "LLM extraction skipped because the current budget/call limit was reached.",
    "llm_extract_skipped_source": "LLM extraction skipped for non-extractable source (search/robots/4xx/etc).",
    "llm_summary_failed": "LLM summary writing failed; pipeline summary fallback used.",
    "field_decision": "Final per-field decision accepted/rejected/unknown.",
    "round_completed": "A run round ended.",
    "run_completed": "Run finished and outputs were written.",
    "max_run_seconds_reached": "Run hit max runtime guard and stopped early."
}

PIPELINE_STAGE_DEFS = [
    ("queued", "Queued", {"queue_transition", "run_started"}),
    ("helper", "Helper Matched", {"helper_files_context_loaded"}),
    ("discover", "Discovery Planned", {"discovery_results_reranked", "source_discovery_only"}),
    ("fetch", "Sources Fetching", {"source_fetch_started"}),
    ("extract", "Extraction", {"source_processed", "field_decision"}),
    ("llm", "LLM Reasoning", {"llm_call_started", "llm_call_completed", "llm_call_usage"}),
    ("round", "Round Complete", {"round_completed"}),
    ("complete", "Run Complete", {"run_completed"})
]


def norm(value) -> str:
    return " ".join(str(value or "").strip().split())


def token(value) -> str:
    return norm(value).lower()


def normalize_field_key(value) -> str:
    text = re.sub(r"[^0-9a-zA-Z]+", "_", norm(value).lower()).strip("_")
    return re.sub(r"_+", "_", text)


def slug(value) -> str:
    return "".join(ch.lower() if ch.isalnum() else "-" for ch in str(value or "")).strip("-").replace("--", "-")


def clean_variant(value) -> str:
    text = norm(value)
    return "" if token(text) in UNKNOWN else text


def known(value) -> bool:
    return token(value) not in UNKNOWN


def is_unknown_token(value) -> bool:
    return token(value) in UNKNOWN


def read_json(path: Path, default=None):
    try:
        with path.open("r", encoding="utf-8") as handle:
            return json.load(handle)
    except Exception:
        return default


def write_json(path: Path, value):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(value, handle, indent=2, ensure_ascii=False)
        handle.write("\n")


def read_jsonl(path: Path, limit: int = 1000):
    if not path.exists():
        return []
    rows = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except Exception:
                continue
    return rows[-limit:]


def read_tail_lines(path: Path, limit: int = 300):
    if not path.exists():
        return []
    try:
        with path.open("r", encoding="utf-8", errors="ignore") as handle:
            lines = handle.readlines()
        return [line.rstrip("\n") for line in lines[-limit:]]
    except Exception:
        return []


def effective_queue_root() -> Path:
    return QUEUE_ROOT


def effective_billing_root() -> Path:
    return BILLING_ROOT


def effective_learning_root() -> Path:
    return LEARNING_ROOT


def helper_category_root(category: str) -> Path:
    return HELPER_ROOT / category


def studio_paths(category: str):
    root = helper_category_root(category)
    control_plane = root / "_control_plane"
    generated = root / "_generated"
    return {
        "root": root,
        "control_plane": control_plane,
        "generated": generated,
        "workbook_map": control_plane / "workbook_map.json",
        "field_rules_draft": control_plane / "field_rules_draft.json",
        "ui_field_catalog_draft": control_plane / "ui_field_catalog_draft.json",
        "field_rules_full": control_plane / "field_rules.full.json",
        "field_rules": generated / "field_rules.json",
        "field_rules_runtime": generated / "field_rules.runtime.json",
        "ui_field_catalog": generated / "ui_field_catalog.json",
        "known_values": generated / "known_values.json",
        "compile_report": generated / "_compile_report.json",
        "tooltip_bank_default": root / "hbs_tooltips.js",
    }


def resolve_local_path(value: str, base: Path | None = None):
    text = str(value or "").strip()
    if not text:
        return None
    candidate = Path(text)
    if candidate.is_absolute():
        return candidate.resolve()
    start = base if base is not None else REPO_ROOT
    return (start / candidate).resolve()


def load_studio_payload(category: str):
    paths = studio_paths(category)
    workbook_map = read_json(paths["workbook_map"], {}) if paths["workbook_map"].exists() else {}
    tooltip_source = workbook_map.get("tooltip_source", {}) if isinstance(workbook_map.get("tooltip_source"), dict) else {}
    tooltip_path_raw = str(tooltip_source.get("path", "") or "").strip()
    default_candidates = [
        paths["root"] / "hbs_tooltips.js",
        paths["root"] / "hbs_tooltipsMouse.js",
    ]
    default_tooltip_path = next((cand for cand in default_candidates if cand.exists()), paths["tooltip_bank_default"])
    tooltip_resolved = resolve_local_path(tooltip_path_raw, paths["root"]) if tooltip_path_raw else default_tooltip_path
    payload = {key: read_json(path, {}) for key, path in paths.items() if key not in {"root", "control_plane", "generated", "tooltip_bank_default", "workbook_map"}}
    payload["workbook_map"] = workbook_map if isinstance(workbook_map, dict) else {}
    payload["paths"] = paths
    payload["tooltip_source_path"] = str(tooltip_resolved) if tooltip_resolved else ""
    payload["tooltip_exists"] = bool(tooltip_resolved and tooltip_resolved.exists())
    return payload


def component_db_snapshot(category: str, component_type: str):
    base = token(component_type)
    if not base:
        return {"entity_count": 0, "sample_entities": []}
    root = helper_category_root(category) / "_generated" / "component_db"
    if not root.exists():
        return {"entity_count": 0, "sample_entities": []}

    canonical_stems = {
        "sensor": ["sensor", "sensors"],
        "switch": ["switch", "switches"],
        "encoder": ["encoder", "encoders"],
        "material": ["material", "materials"],
        "mcu": ["mcu", "mcus"],
    }
    stems = canonical_stems.get(base, [base])
    if base not in canonical_stems:
        if base.endswith("y") and len(base) > 1:
            stems.append(f"{base[:-1]}ies")
        stems.append(f"{base}s")
        stems.append(f"{base}es")

    deduped_stems = []
    seen = set()
    for stem in stems:
        s = str(stem or "").strip().lower()
        if s and s not in seen:
            seen.add(s)
            deduped_stems.append(s)

    for stem in deduped_stems:
        payload = read_json(root / f"{stem}.json", {})
        if not isinstance(payload, dict):
            continue
        items = payload.get("items", [])
        if not isinstance(items, list):
            continue
        sample_entities = []
        for item in items:
            if isinstance(item, dict):
                name = str(item.get("name", item.get("id", "")) or "").strip()
            else:
                name = str(item or "").strip()
            if not name:
                continue
            sample_entities.append(name)
            if len(sample_entities) >= 10:
                break
        return {
            "entity_count": len(items),
            "sample_entities": sample_entities,
        }

    return {"entity_count": 0, "sample_entities": []}


def upsert_ui_catalog_row(ui_draft: dict, key: str, ui_data: dict, rule: dict):
    fields = ui_draft.setdefault("fields", [])
    if not isinstance(fields, list):
        fields = []
        ui_draft["fields"] = fields
    row = next((r for r in fields if isinstance(r, dict) and r.get("key") == key), None)
    if row is None:
        row = {"key": key}
        fields.append(row)
    priority = rule_priority(rule)
    contract = rule_contract(rule)
    row.update(
        {
            "key": key,
            "canonical_key": rule.get("canonical_key") or key,
            "label": ui_data.get("label", key),
            "short_label": ui_data.get("short_label"),
            "group": ui_data.get("group", "General"),
            "order": int(ui_data.get("order", 1) or 1),
            "tooltip_md": ui_data.get("tooltip_md", ""),
            "tooltip_key": ui_data.get("tooltip_key"),
            "tooltip_source": ui_data.get("tooltip_source"),
            "prefix": ui_data.get("prefix"),
            "suffix": ui_data.get("suffix"),
            "placeholder": ui_data.get("placeholder", "unk"),
            "input_control": ui_data.get("input_control", "text"),
            "display_mode": ui_data.get("display_mode", "all"),
            "display_decimals": int(ui_data.get("display_decimals", 0) or 0),
            "array_handling": rule.get("array_handling", ui_data.get("array_handling", "none")),
            "aliases": rule.get("aliases", []),
            "examples": ui_data.get("examples", []),
            "required_level": priority.get("required_level", "optional") or "optional",
            "availability": priority.get("availability", "sometimes") or "sometimes",
            "difficulty": priority.get("difficulty", "medium") or "medium",
            "effort": int(priority.get("effort", 5) or 5),
            "type": contract.get("type", "string") or "string",
            "shape": contract.get("shape", "scalar") or "scalar",
            "unit": contract.get("unit", ""),
        }
    )


def save_studio_drafts(category: str, field_rules_draft: dict, ui_field_catalog_draft: dict):
    paths = studio_paths(category)
    write_json(paths["field_rules_draft"], field_rules_draft)
    write_json(paths["ui_field_catalog_draft"], ui_field_catalog_draft)


def csv_tokens(value: str):
    if value is None:
        return []
    raw = str(value).replace("\n", ",")
    return [item.strip() for item in raw.split(",") if item.strip()]


def csv_join(values):
    return ", ".join([str(v).strip() for v in (values or []) if str(v).strip()])


def parse_key_value_lines(text: str):
    out = {}
    for line in str(text or "").splitlines():
        row = line.strip()
        if not row or "=" not in row:
            continue
        key, value = row.split("=", 1)
        key = key.strip()
        value = value.strip()
        if not key:
            continue
        try:
            out[key] = float(value) if "." in value else int(value)
        except Exception:
            out[key] = value
    return out


def format_key_value_lines(payload: dict):
    if not isinstance(payload, dict):
        return ""
    lines = []
    for key in sorted(payload.keys()):
        lines.append(f"{key}={payload[key]}")
    return "\n".join(lines)


def collect_sheet_names(workbook_map: dict):
    names = set()
    if not isinstance(workbook_map, dict):
        return []
    key_list = workbook_map.get("key_list", {}) if isinstance(workbook_map.get("key_list"), dict) else {}
    product_table = workbook_map.get("product_table", {}) if isinstance(workbook_map.get("product_table"), dict) else {}
    if key_list.get("sheet"):
        names.add(str(key_list["sheet"]))
    if product_table.get("sheet"):
        names.add(str(product_table["sheet"]))
    for row in workbook_map.get("sheet_roles", []) or []:
        if isinstance(row, dict) and row.get("sheet"):
            names.add(str(row["sheet"]))
    for row in workbook_map.get("enum_lists", []) or []:
        if isinstance(row, dict) and row.get("sheet"):
            names.add(str(row["sheet"]))
    for row in workbook_map.get("component_sheets", []) or []:
        if isinstance(row, dict) and row.get("sheet"):
            names.add(str(row["sheet"]))
    for row in workbook_map.get("component_sources", []) or []:
        if isinstance(row, dict) and row.get("sheet"):
            names.add(str(row["sheet"]))
    return sorted([name for name in names if name.strip()])


def col_to_index(column: str):
    token_value = str(column or "").strip().upper()
    if not token_value:
        return None
    total = 0
    for ch in token_value:
        if ch < "A" or ch > "Z":
            return None
        total = (total * 26) + (ord(ch) - 64)
    return total if total > 0 else None


def index_to_col(index: int):
    value = int(index or 0)
    if value <= 0:
        return ""
    out = ""
    while value > 0:
        rem = (value - 1) % 26
        out = chr(65 + rem) + out
        value = (value - 1) // 26
    return out


def normalize_col(value, fallback: str = ""):
    token_value = str(value or "").strip().upper()
    return token_value if col_to_index(token_value) else str(fallback or "").strip().upper()


def stable_sort_columns(values):
    uniq = []
    seen = set()
    for value in values or []:
        col = normalize_col(value)
        if not col or col in seen:
            continue
        seen.add(col)
        uniq.append(col)
    uniq.sort(key=lambda col: col_to_index(col) or 9999)
    return uniq


def normalize_component_rows(workbook_map: dict):
    rows = []
    rows_raw = workbook_map.get("component_sheets", []) if isinstance(workbook_map.get("component_sheets"), list) else []
    if not rows_raw:
        rows_raw = workbook_map.get("component_sources", []) if isinstance(workbook_map.get("component_sources"), list) else []
    for row in rows_raw:
        if not isinstance(row, dict):
            continue
        header_row = int(row.get("header_row", 1) or 1)
        first_data_row = int(row.get("first_data_row", row.get("row_start", row.get("start_row", header_row + 1)) or (header_row + 1)) or (header_row + 1))
        stop_after_blank_names = int(row.get("stop_after_blank_names", 10) or 10)
        rows.append(
            {
                "sheet": str(row.get("sheet", "") or "").strip(),
                "component_type": str(row.get("component_type", row.get("type", "")) or "").strip(),
                "header_row": max(1, header_row),
                "first_data_row": max(1, first_data_row),
                "canonical_name_column": normalize_col(row.get("canonical_name_column", row.get("name_column", "A")), "A"),
                "brand_column": normalize_col(row.get("brand_column", ""), ""),
                "alias_columns": stable_sort_columns(row.get("alias_columns", [])),
                "link_columns": stable_sort_columns(row.get("link_columns", [])),
                "property_columns": stable_sort_columns(row.get("property_columns", [])),
                "auto_derive_aliases": bool(row.get("auto_derive_aliases", True)),
                "stop_after_blank_names": max(1, stop_after_blank_names),
            }
        )
    if not rows:
        rows = [
            {
                "sheet": "",
                "component_type": "",
                "header_row": 1,
                "first_data_row": 2,
                "canonical_name_column": "A",
                "brand_column": "",
                "alias_columns": [],
                "link_columns": [],
                "property_columns": [],
                "auto_derive_aliases": True,
                "stop_after_blank_names": 10,
            }
        ]
    return rows


def _resolve_workbook_path(value: str):
    text = str(value or "").strip()
    if not text:
        return None
    path = Path(text)
    if not path.is_absolute():
        path = (REPO_ROOT / path).resolve()
    return path


def ordered_unique_text(values):
    out = []
    seen = set()
    for raw in values or []:
        text = norm(raw)
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(text)
    return out


def split_component_tokens(value):
    text = norm(value)
    if not text:
        return []
    parts = [norm(part) for part in re.split(r"[,\n;|/]+", text)]
    return [part for part in parts if part]


def extract_http_links(value):
    text = str(value or "").strip()
    if not text:
        return []
    matches = [norm(link) for link in re.findall(r"https?://[^\s,;|]+", text, flags=re.IGNORECASE)]
    if matches:
        return ordered_unique_text(matches)
    normalized = norm(text)
    if normalized.lower().startswith("http://") or normalized.lower().startswith("https://"):
        return [normalized]
    return []


def _toggle_csv_column(raw_csv: str, col: str):
    values = stable_sort_columns(csv_tokens(raw_csv))
    target = normalize_col(col, "")
    if not target:
        return csv_join(values)
    if target in values:
        values = [value for value in values if value != target]
    else:
        values.append(target)
    return csv_join(stable_sort_columns(values))


def derive_safe_aliases(name: str):
    base = norm(name)
    if not base:
        return []
    variants = []
    spaced = norm(re.sub(r"[^0-9A-Za-z]+", " ", base))
    compact = norm(re.sub(r"[^0-9A-Za-z]+", "", base))
    if spaced and token(spaced) != token(base):
        variants.append(spaced)
    if compact and len(compact) >= 4 and token(compact) not in {token(base), token(spaced)}:
        variants.append(compact)
    return ordered_unique_text([value for value in variants if token(value) != token(base)])


@st.cache_data(ttl=20, show_spinner=False)
def workbook_component_preview(workbook_path: str, sheet_name: str, header_row: int, first_data_row: int):
    if load_workbook is None:
        return {"error": "openpyxl is not installed in this environment."}
    path = _resolve_workbook_path(workbook_path)
    if not path or not path.exists():
        return {"error": f"Workbook not found: {workbook_path}"}
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return {"error": f"Failed to read workbook: {exc}"}
    if str(sheet_name or "").strip() not in wb.sheetnames:
        return {"error": f"Sheet not found: {sheet_name}"}
    ws = wb[str(sheet_name)]
    max_scan_col = max(1, min(int(ws.max_column or 1), 32))
    columns = []
    for col_idx in range(1, max_scan_col + 1):
        col = index_to_col(col_idx)
        header = norm(ws.cell(row=max(1, int(header_row or 1)), column=col_idx).value)
        raw_values = []
        for row_idx in range(max(1, int(first_data_row or 2)), max(1, int(first_data_row or 2)) + 20):
            raw_values.append(norm(ws.cell(row=row_idx, column=col_idx).value))
        if not header and not any(raw_values):
            continue
        columns.append(
            {
                "col": col,
                "header": header,
                "values": raw_values,
            }
        )
    preview_rows = []
    preview_cols = columns[:10]
    for offset in range(0, 8):
        row_idx = max(1, int(first_data_row or 2)) + offset
        row = {"row": row_idx}
        for col in preview_cols:
            key = f"{col['col']}:{col['header'] or '(blank)'}"
            row[key] = col["values"][offset] if offset < len(col["values"]) else ""
        preview_rows.append(row)
    return {"columns": columns, "preview_rows": preview_rows}


@st.cache_data(ttl=20, show_spinner=False)
def workbook_component_entities_preview(
    workbook_path: str,
    sheet_name: str,
    header_row: int,
    first_data_row: int,
    canonical_name_column: str,
    brand_column: str = "",
    alias_columns_csv: str = "",
    link_columns_csv: str = "",
    auto_derive_aliases: bool = True,
    stop_after_blank_names: int = 10,
):
    if load_workbook is None:
        return {"error": "openpyxl is not installed in this environment."}
    path = _resolve_workbook_path(workbook_path)
    if not path or not path.exists():
        return {"error": f"Workbook not found: {workbook_path}"}
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return {"error": f"Failed to read workbook: {exc}"}
    if str(sheet_name or "").strip() not in wb.sheetnames:
        return {"error": f"Sheet not found: {sheet_name}"}

    ws = wb[str(sheet_name)]
    name_col = normalize_col(canonical_name_column, "A")
    brand_col = normalize_col(brand_column, "")
    alias_cols = stable_sort_columns(csv_tokens(alias_columns_csv))
    link_cols = stable_sort_columns(csv_tokens(link_columns_csv))
    stop_limit = max(1, int(stop_after_blank_names or 10))
    row_end = int(ws.max_row or max(2, int(first_data_row or 2)))

    preview_rows = []
    sample_names = []
    sample_name_tokens = set()
    first_20_names = []
    blank_streak = 0
    scanned_rows = 0
    entity_count = 0
    numeric_only_names = 0

    for row_idx in range(max(1, int(first_data_row or 2)), row_end + 1):
        scanned_rows += 1
        name = norm(ws.cell(row=row_idx, column=col_to_index(name_col) or 1).value)
        if not name:
            blank_streak += 1
            if blank_streak >= stop_limit:
                break
            continue
        blank_streak = 0
        entity_count += 1
        if re.match(r"^\d+$", name):
            numeric_only_names += 1
        if len(first_20_names) < 20:
            first_20_names.append(name)
        name_token = token(name)
        if len(sample_names) < 10 and name_token not in sample_name_tokens:
            sample_name_tokens.add(name_token)
            sample_names.append(name)

        brand = ""
        if brand_col:
            brand = norm(ws.cell(row=row_idx, column=col_to_index(brand_col) or 1).value)

        alias_values = []
        for col in alias_cols:
            alias_values.extend(split_component_tokens(ws.cell(row=row_idx, column=col_to_index(col) or 1).value))
        if not alias_values and bool(auto_derive_aliases):
            alias_values.extend(derive_safe_aliases(name))
        aliases = [value for value in ordered_unique_text(alias_values) if token(value) != token(name)]

        link_values = []
        for col in link_cols:
            link_values.extend(extract_http_links(ws.cell(row=row_idx, column=col_to_index(col) or 1).value))
        links = ordered_unique_text(link_values)

        if len(preview_rows) < 20:
            preview_rows.append(
                {
                    "row": row_idx,
                    "name": name,
                    "brand": brand,
                    "aliases": aliases,
                    "links": links,
                }
            )

    numeric_ratio = (numeric_only_names / entity_count) if entity_count > 0 else 0.0
    return {
        "preview_rows": preview_rows,
        "entity_count": int(entity_count),
        "sample_names": sample_names[:10],
        "first_20_names": first_20_names,
        "numeric_only_names": int(numeric_only_names),
        "numeric_only_ratio": float(numeric_ratio),
        "first_20_all_numeric": bool(first_20_names and all(re.match(r"^\d+$", value) for value in first_20_names)),
        "scanned_rows": int(scanned_rows),
    }


@st.cache_data(ttl=20, show_spinner=False)
def workbook_sheet_names(workbook_path: str):
    if load_workbook is None:
        return []
    path = _resolve_workbook_path(workbook_path)
    if not path or not path.exists():
        return []
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return []
    return [str(name) for name in wb.sheetnames if str(name or "").strip()]


def numeric_only_ratio(values):
    cleaned = [str(v or "").strip() for v in values or [] if str(v or "").strip()]
    if not cleaned:
        return 0.0
    count = sum(1 for value in cleaned if re.match(r"^\d+$", value))
    return count / len(cleaned)


def component_preview_errors(name_values):
    cleaned = [str(v or "").strip() for v in (name_values or []) if str(v or "").strip()]
    if not cleaned:
        return ["No non-empty values found in selected Name/Model column preview."]
    ratio = numeric_only_ratio(cleaned)
    if ratio > 0.10:
        return ["More than 10% of preview values are numeric-only. Choose the Name/Model column, not the ID column."]
    if cleaned and all(re.match(r"^\d+$", value) for value in cleaned[:20]):
        return ["The first 20 preview values are numeric-only. Choose the Name/Model column."]
    return []


def collect_unique_samples(raw_values, include_unknown: bool = False, max_unique: int = 10, max_scan: int = 100, max_blank_streak: int = 20):
    uniques = []
    seen = set()
    scanned = 0
    blank_streak = 0
    for raw in raw_values or []:
        if scanned >= max(1, int(max_scan or 100)):
            break
        scanned += 1
        text = norm(raw)
        if not text or (is_unknown_token(text) and not include_unknown):
            blank_streak += 1
            if blank_streak >= max(1, int(max_blank_streak or 20)):
                break
            continue
        blank_streak = 0
        bucket = token(text)
        if bucket in seen:
            continue
        seen.add(bucket)
        uniques.append(text)
        if len(uniques) >= max(1, int(max_unique or 10)):
            break
    return {
        "samples": uniques,
        "scanned": scanned,
        "blank_streak": blank_streak,
    }


def find_key_row_hint(workbook_map: dict, selected_key: str, fallback_row: int = 0):
    if int(fallback_row or 0) > 0:
        return int(fallback_row)
    key_map = workbook_map.get("key_list", {}) if isinstance(workbook_map.get("key_list"), dict) else {}
    key_sheet = str(key_map.get("sheet", "") or "").strip()
    key_column = normalize_col(key_map.get("column", "B"), "B")
    row_start = int(key_map.get("row_start", 1) or 1)
    row_end = int(key_map.get("row_end", 0) or 0)
    workbook_path = str(workbook_map.get("workbook_path", "") or "").strip()
    if load_workbook is None or not workbook_path or not key_sheet:
        return 0
    path = _resolve_workbook_path(workbook_path)
    if not path or not path.exists():
        return 0
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return 0
    if key_sheet not in wb.sheetnames:
        return 0
    ws = wb[key_sheet]
    end_row = row_end if row_end > 0 else int(ws.max_row or row_start)
    target = normalize_field_key(selected_key)
    col_idx = col_to_index(key_column) or 2
    for row_idx in range(max(1, row_start), max(1, end_row) + 1):
        candidate = normalize_field_key(ws.cell(row=row_idx, column=col_idx).value)
        if candidate == target:
            return row_idx
    return 0


@st.cache_data(ttl=20, show_spinner=False)
def workbook_context_samples(
    workbook_path: str,
    workbook_map_json: str,
    selected_key: str,
    key_row_hint: int,
    include_unknown: bool = False,
    max_unique: int = 10,
    max_scan: int = 100,
    max_blank_streak: int = 20,
):
    if load_workbook is None:
        return {"error": "openpyxl is not installed in this environment.", "samples": []}
    path = _resolve_workbook_path(workbook_path)
    if not path or not path.exists():
        return {"error": f"Workbook not found: {workbook_path}", "samples": []}
    try:
        workbook_map = json.loads(str(workbook_map_json or "{}"))
    except Exception:
        workbook_map = {}
    product_map = workbook_map.get("product_table", {}) if isinstance(workbook_map.get("product_table"), dict) else {}
    layout = str(product_map.get("layout", "matrix") or "matrix").strip().lower()
    sample_sheet_name = str(product_map.get("sheet", "") or "").strip()
    if not sample_sheet_name:
        return {"error": "Sampling sheet is not configured in workbook_map.product_table.sheet.", "samples": []}
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return {"error": f"Failed to read workbook: {exc}", "samples": []}
    if sample_sheet_name not in wb.sheetnames:
        return {"error": f"Sampling sheet not found: {sample_sheet_name}", "samples": []}
    ws = wb[sample_sheet_name]

    selected_key_norm = normalize_field_key(selected_key)
    raw_values = []

    if layout in {"matrix", "column_range"}:
        row_idx = int(key_row_hint or 0)
        if row_idx <= 0:
            return {"error": f"Unable to locate row for key `{selected_key_norm}`.", "samples": []}
        start_col = col_to_index(product_map.get("value_col_start", "C")) or 3
        end_col = col_to_index(product_map.get("value_col_end", "")) or int(ws.max_column or start_col)
        end_col = max(start_col, end_col)
        brand_row = int(product_map.get("brand_row", 3) or 3)
        model_row = int(product_map.get("model_row", 4) or 4)
        for col_idx in range(start_col, end_col + 1):
            brand = norm(ws.cell(row=brand_row, column=col_idx).value)
            model = norm(ws.cell(row=model_row, column=col_idx).value)
            if not brand and not model:
                continue
            raw_values.append(ws.cell(row=row_idx, column=col_idx).value)
    elif layout in {"row_table", "rows"}:
        header_row = int(product_map.get("header_row", 1) or 1)
        data_row_start = int(product_map.get("data_row_start", header_row + 1) or (header_row + 1))
        col_idx = 0
        for idx in range(1, int(ws.max_column or 1) + 1):
            header_value = normalize_field_key(ws.cell(row=header_row, column=idx).value)
            if header_value == selected_key_norm:
                col_idx = idx
                break
        if col_idx <= 0:
            return {"error": f"Key column not found for `{selected_key_norm}` in `{sample_sheet_name}` row {header_row}.", "samples": []}
        for row_idx in range(data_row_start, int(ws.max_row or data_row_start) + 1):
            raw_values.append(ws.cell(row=row_idx, column=col_idx).value)
    else:
        return {"error": f"Unsupported sampling layout `{layout}`.", "samples": []}

    sampled = collect_unique_samples(
        raw_values=raw_values,
        include_unknown=bool(include_unknown),
        max_unique=int(max_unique or 10),
        max_scan=int(max_scan or 100),
        max_blank_streak=int(max_blank_streak or 20),
    )
    return {
        "samples": sampled.get("samples", []),
        "scanned": int(sampled.get("scanned", 0) or 0),
        "blank_streak": int(sampled.get("blank_streak", 0) or 0),
        "layout": layout,
        "sheet": sample_sheet_name,
    }


def _first_text(*values):
    for value in values:
        text = str(value or "").strip()
        if text:
            return text
    return ""


def _first_defined(*values):
    for value in values:
        if value is not None:
            return value
    return None


def rule_priority(rule: dict):
    if not isinstance(rule, dict):
        return {}
    base = rule.get("priority", {}) if isinstance(rule.get("priority"), dict) else {}
    effort_raw = _first_defined(rule.get("effort"), base.get("effort"), 0)
    publish_gate_raw = _first_defined(rule.get("publish_gate"), base.get("publish_gate"), False)
    block_publish_when_unk_raw = _first_defined(
        rule.get("block_publish_when_unk"),
        base.get("block_publish_when_unk"),
        False,
    )
    return {
        "required_level": _first_text(rule.get("required_level"), base.get("required_level")),
        "availability": _first_text(rule.get("availability"), base.get("availability")),
        "difficulty": _first_text(rule.get("difficulty"), base.get("difficulty")),
        "effort": int(effort_raw or 0),
        "publish_gate": bool(publish_gate_raw),
        "block_publish_when_unk": bool(block_publish_when_unk_raw),
    }


def rule_contract(rule: dict):
    if not isinstance(rule, dict):
        return {}
    base = rule.get("contract", {}) if isinstance(rule.get("contract"), dict) else {}
    type_value = _first_text(rule.get("type"), base.get("type"))
    shape_value = _first_text(rule.get("shape"), base.get("shape"))
    unit_value = _first_text(rule.get("unit"), base.get("unit"))
    return {
        **base,
        "type": type_value,
        "shape": shape_value,
        "unit": unit_value,
    }


def rule_parse(rule: dict):
    if not isinstance(rule, dict):
        return {}
    base = rule.get("parse", {}) if isinstance(rule.get("parse"), dict) else {}
    parse_template = _first_text(rule.get("parse_template"), base.get("template"))
    out = dict(base)
    out["template"] = parse_template
    if "strict_unit_required" in rule and rule.get("strict_unit_required") is not None:
        out["strict_unit_required"] = bool(rule.get("strict_unit_required"))
    return out


def rule_enum(rule: dict):
    if not isinstance(rule, dict):
        return {}
    base = rule.get("enum", {}) if isinstance(rule.get("enum"), dict) else {}
    enum_policy = _first_text(rule.get("enum_policy"), base.get("policy"))
    source = rule.get("enum_source")
    if normalize_enum_source_value(source) == "none":
        source = base.get("source")
    return {
        **base,
        "policy": enum_policy,
        "source": source,
    }


def rule_min_evidence(rule: dict):
    if not isinstance(rule, dict):
        return 1
    evidence = rule.get("evidence", {}) if isinstance(rule.get("evidence"), dict) else {}
    value = _first_defined(rule.get("min_evidence_refs"), evidence.get("min_evidence_refs"), 1)
    return int(value or 1)


def key_contract_errors(rule: dict):
    errors = []
    if not isinstance(rule, dict):
        return ["Missing rule payload."]
    priority = rule_priority(rule)
    contract = rule_contract(rule)
    parse = rule_parse(rule)
    enum_obj = rule_enum(rule)
    ui = rule.get("ui", {}) if isinstance(rule.get("ui"), dict) else {}
    if not str(ui.get("label", "")).strip():
        errors.append("Missing label")
    if not str(ui.get("group", "")).strip():
        errors.append("Missing group")
    if int(ui.get("order", 0) or 0) <= 0:
        errors.append("Missing/invalid order")
    has_tooltip_key = str(ui.get("tooltip_key", "")).strip()
    has_tooltip_md = "tooltip_md" in ui
    if not has_tooltip_key and not has_tooltip_md:
        errors.append("Missing tooltip_key or tooltip_md")
    if not str(priority.get("required_level", "optional") or "optional").strip():
        errors.append("Missing required_level")
    if not str(priority.get("availability", "sometimes") or "sometimes").strip():
        errors.append("Missing availability")
    if not str(priority.get("difficulty", "medium") or "medium").strip():
        errors.append("Missing difficulty")
    effort = int(priority.get("effort", 0) or 0)
    if effort < 1 or effort > 10:
        errors.append("Effort must be 1..10")
    if not str(contract.get("type", "")).strip():
        errors.append("Missing type")
    if not str(contract.get("shape", "")).strip():
        errors.append("Missing shape")
    list_rules = contract.get("list_rules", {}) if isinstance(contract.get("list_rules"), dict) else {}
    if not list_rules and isinstance(rule.get("list_rules"), dict):
        list_rules = rule.get("list_rules")
    if str(contract.get("shape", "")).strip() == "list" and not isinstance(list_rules, dict):
        errors.append("List shape requires list_rules")
    if str(contract.get("shape", "")).strip() == "list" and isinstance(list_rules, dict) and not list_rules:
        errors.append("List shape requires list_rules")
    object_schema = contract.get("object_schema", {}) if isinstance(contract.get("object_schema"), dict) else {}
    if not object_schema and isinstance(rule.get("object_schema"), dict):
        object_schema = rule.get("object_schema")
    if str(contract.get("shape", "")).strip() == "object" and not object_schema:
        errors.append("Object shape requires object_schema")
    parse_rules = rule.get("parse_rules", {}) if isinstance(rule.get("parse_rules"), dict) else {}

    rule_type = str(contract.get("type", "")).strip()
    rule_shape = str(contract.get("shape", "")).strip()
    unit_value = str(contract.get("unit", "")).strip()
    if rule_type in {"number", "integer"} and not unit_value:
        errors.append("Numeric field missing unit")

    strict_unit = (
        rule.get("strict_unit_required")
        if "strict_unit_required" in rule
        else parse.get("strict_unit_required")
    )
    unit_accepts = parse.get("unit_accepts")
    if unit_accepts is None:
        unit_accepts = parse_rules.get("unit_accepts")
    if rule_type in {"number", "integer"} and isinstance(unit_accepts, list) and len(unit_accepts) > 1 and strict_unit is None:
        errors.append("Numeric field with conversions requires strict_unit_required decision")

    enum_policy = str(enum_obj.get("policy", "")).strip()
    enum_source_token = normalize_enum_source_value(
        enum_obj.get("source")
    )
    if enum_policy in {"closed", "closed_with_curation"} and enum_source_token == "none":
        errors.append("Closed enum policy requires enum_source")
    if enum_policy in {"open", "open_prefer_known"}:
        nvp = rule.get("new_value_policy", {}) if isinstance(rule.get("new_value_policy"), dict) else {}
        if not nvp and isinstance(enum_obj.get("new_value_policy"), dict):
            nvp = enum_obj.get("new_value_policy")
        if not nvp:
            errors.append("Open enum policy requires new_value_policy")
        elif "accept_if_evidence" not in nvp or "mark_needs_curation" not in nvp:
            errors.append("Open enum policy new_value_policy must include accept_if_evidence + mark_needs_curation")

    value_form = str(rule.get("value_form", "")).strip().lower()
    if rule_type in {"number", "integer"} and rule_shape in {"list", "range", "object"} and not value_form:
        errors.append("Numeric list/range field requires value_form")

    parse_template = str(parse.get("template", "")).strip()
    list_templates = {
        "list_of_tokens_delimited",
        "list_of_numbers_with_unit",
        "mode_tagged_list",
        "mode_tagged_values",
        "latency_list_modes_ms",
        "list_numbers_or_ranges_with_unit",
    }
    if parse_template in list_templates and rule_shape != "list":
        errors.append(f"{parse_template} requires shape=list")
    if parse_template == "component_reference" and not str(enum_source_token).startswith("component_db:"):
        component_block = rule.get("component", {}) if isinstance(rule.get("component"), dict) else {}
        component_source = str(component_block.get("source", "") or "")
        if not (
            component_source.startswith("component_db.")
            or component_source.startswith("component_db:")
            or component_source.startswith("component_db_sources:")
        ):
            errors.append("component_reference requires component_db source")
    if parse_template == "latency_list_modes_ms":
        object_schema = rule.get("object_schema")
        if not isinstance(object_schema, dict):
            object_schema = contract.get("object_schema", {})
        if not isinstance(object_schema, dict) or not object_schema:
            errors.append("latency_list_modes_ms requires object_schema")
    return errors


def normalize_enum_source_value(enum_source):
    if isinstance(enum_source, str):
        source_text = str(enum_source).strip()
        if source_text and "." not in source_text and ":" not in source_text:
            return f"enum_buckets:{source_text}"
        if "." in source_text:
            source_type, source_ref = source_text.split(".", 1)
            if source_type.strip() and source_ref.strip():
                source_type = source_type.strip()
                source_ref = source_ref.strip()
                if source_type in {"known_values", "enum_buckets", "data_lists", "datalists"}:
                    return f"enum_buckets:{source_ref}"
                return f"{source_type}:{source_ref}"
        if ":" in source_text:
            source_type, source_ref = source_text.split(":", 1)
            if source_type.strip() and source_ref.strip():
                source_type = source_type.strip()
                source_ref = source_ref.strip()
                if source_type in {"known_values", "enum_buckets", "data_lists", "datalists"}:
                    return f"enum_buckets:{source_ref}"
                return f"{source_type}:{source_ref}"
        return "none"
    if isinstance(enum_source, dict):
        source_type = str(enum_source.get("type", "")).strip()
        source_ref = str(enum_source.get("ref", "")).strip()
        if source_type and source_ref:
            if source_type in {"known_values", "enum_buckets", "data_lists", "datalists"}:
                return f"enum_buckets:{source_ref}"
            return f"{source_type}:{source_ref}"
    return "none"


def apply_enum_source_value(selected: str):
    token_value = str(selected or "").strip()
    if token_value in {"", "none"} or ":" not in token_value:
        return None
    source_type, source_ref = token_value.split(":", 1)
    source_type = source_type.strip()
    source_ref = source_ref.strip()
    if not source_type or not source_ref:
        return None
    if source_type in {"enum_buckets", "known_values", "data_lists", "datalists"}:
        source_type = "known_values"
    return {"type": source_type, "ref": source_ref}


def enum_source_ref_to_string(enum_source: dict | None):
    if not isinstance(enum_source, dict):
        return None
    source_type = str(enum_source.get("type", "")).strip().lower()
    source_ref = str(enum_source.get("ref", "")).strip()
    if not source_type or not source_ref:
        return None
    if source_type in {"known_values", "enum_buckets", "data_lists", "datalists"}:
        return f"enum_buckets:{source_ref}"
    return f"{source_type}.{source_ref}"


@st.cache_data(ttl=30, show_spinner=False)
def load_tooltip_bank_entries(tooltip_path: str):
    path = resolve_local_path(str(tooltip_path or ""))
    if not path or not path.exists():
        return {}
    try:
        text = path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return {}

    suffix = path.suffix.lower().strip()
    out = {}

    def push_entry(raw_key: str, html: str = "", markdown: str = ""):
        key = normalize_field_key(raw_key)
        if not key:
            return
        html_text = str(html or "").strip()
        markdown_text = str(markdown or "").strip()
        if not markdown_text and html_text:
            markdown_text = re.sub(r"<[^>]+>", " ", html_text)
            markdown_text = re.sub(r"\s+", " ", markdown_text).strip()
        plain = re.sub(r"<[^>]+>", " ", markdown_text or html_text)
        plain = re.sub(r"\s+", " ", plain).strip()
        if not (html_text or markdown_text or plain):
            return
        out[key] = {
            "key": key,
            "html": html_text,
            "markdown": markdown_text or plain,
            "plain": plain,
            "source": path.name,
        }

    if suffix == ".json":
        try:
            payload = json.loads(text)
        except Exception:
            payload = {}
        bucket = payload.get("tooltips") if isinstance(payload, dict) and isinstance(payload.get("tooltips"), dict) else payload
        if isinstance(bucket, dict):
            for raw_key, value in bucket.items():
                if isinstance(value, str):
                    push_entry(str(raw_key), markdown=value)
                elif isinstance(value, dict):
                    push_entry(
                        str(raw_key),
                        html=str(value.get("html", "") or ""),
                        markdown=str(
                            value.get("markdown", "")
                            or value.get("md", "")
                            or value.get("tooltip_md", "")
                            or value.get("text", "")
                            or ""
                        ),
                    )
        return out

    if suffix in {".md", ".markdown"}:
        current_key = ""
        buffer = []
        for line in text.splitlines():
            heading = re.match(r"^\s{0,3}#{1,6}\s+(.+?)\s*$", line or "")
            if heading:
                if current_key and buffer:
                    push_entry(current_key, markdown="\n".join(buffer).strip())
                current_key = normalize_field_key(re.sub(r"`", "", str(heading.group(1) or "").strip()))
                buffer = []
                continue
            if current_key:
                buffer.append(line)
        if current_key and buffer:
            push_entry(current_key, markdown="\n".join(buffer).strip())
        return out

    # Default parser for JS/TS and loose text tooltip banks.
    js_template_pattern = re.compile(r'([A-Za-z0-9_]+)\s*:\s*`([\s\S]*?)`\s*(?:,|$)', re.MULTILINE)
    for match in js_template_pattern.finditer(text):
        push_entry(str(match.group(1) or ""), html=str(match.group(2) or ""))
    if out:
        return out

    js_string_pattern = re.compile(r'["\']([A-Za-z0-9_\- ]+)["\']\s*:\s*["\']([^"\']+)["\']\s*(?:,|$)', re.MULTILINE)
    for match in js_string_pattern.finditer(text):
        push_entry(str(match.group(1) or ""), markdown=str(match.group(2) or ""))
    return out


def try_normalize_preview(raw_value: str, parse_template: str, unit_conversions: dict, delimiters):
    value = str(raw_value or "").strip()
    if not value:
        return "unk"
    template = str(parse_template or "").strip()
    if template == "boolean_yes_no_unknown":
        token_value = value.lower()
        if token_value in {"yes", "true", "1", "on"}:
            return "yes"
        if token_value in {"no", "false", "0", "off"}:
            return "no"
        return "unk"
    if template in {"list_of_tokens_delimited", "mode_tagged_values"}:
        d = delimiters or [","]
        pattern = "|".join(re.escape(item) for item in d if item) or ","
        parts = [part.strip() for part in re.split(pattern, value) if part.strip()]
        return parts or "unk"
    if template in {"number_with_unit", "integer_with_unit", "range_number", "list_numbers_or_ranges_with_unit"}:
        match = re.search(r"-?\d+(?:\.\d+)?", value)
        if not match:
            return "unk"
        number_value = float(match.group(0))
        lowered = value.lower()
        for conv_key, conv_value in (unit_conversions or {}).items():
            key = str(conv_key).lower().strip()
            if "_to_" not in key:
                continue
            src_unit = key.split("_to_", 1)[0]
            if src_unit and src_unit in lowered:
                try:
                    number_value = number_value * float(conv_value)
                except Exception:
                    pass
                break
        if template == "integer_with_unit":
            return int(round(number_value))
        return number_value
    return value


def infer_unit_hint(field_key: str) -> str:
    token_value = token(field_key)
    if token_value in {"weight"} or token_value.endswith("_weight"):
        return "g"
    if token_value in {"dpi"} or token_value.endswith("_dpi"):
        return "dpi"
    if "polling" in token_value or token_value.endswith("_hz"):
        return "hz"
    if token_value in {"lngth", "length", "width", "height"} or token_value.endswith("_length") or token_value.endswith("_width") or token_value.endswith("_height"):
        return "mm"
    if "battery" in token_value and "hour" in token_value:
        return "h"
    if "latency" in token_value:
        return "ms"
    return ""


def infer_parse_template_hint(field_key: str, inferred_type: str, inferred_shape: str, component_hint: bool = False) -> str:
    key_token = token(field_key)
    if component_hint or any(tag in key_token for tag in ["sensor", "switch", "encoder", "material", "mcu"]):
        return "component_reference"
    if inferred_type == "boolean":
        return "boolean_yes_no_unknown"
    if inferred_shape == "list" and inferred_type in {"number", "integer"}:
        if "latency" in key_token:
            return "latency_list_modes_ms"
        if "lift" in key_token:
            return "list_numbers_or_ranges_with_unit"
        return "list_of_numbers_with_unit"
    if inferred_shape == "list":
        if any(tag in key_token for tag in ["connection", "connectivity", "wired", "wireless", "bluetooth"]):
            return "mode_tagged_values"
        return "list_of_tokens_delimited"
    if inferred_type in {"number", "integer"}:
        return "number_with_unit"
    if inferred_type == "date":
        return "date_field"
    if inferred_type == "url":
        return "url_field"
    return "text_field"


def infer_key_hints(selected_key: str, sample_values):
    key_token = token(selected_key)
    values = [str(v).strip() for v in (sample_values or []) if str(v).strip()]
    numeric = 0
    numeric_or_unit = 0
    bool_like = 0
    list_like = 0
    range_like = 0
    url_like = 0
    date_like = 0
    for raw in values:
        low = raw.lower()
        if re.search(r"^-?\d+(?:\.\d+)?$", low):
            numeric += 1
        if re.search(r"-?\d+(?:\.\d+)?\s*(mm|cm|in|\"|g|kg|hz|khz|dpi|ips|ms|s|sec|seconds|h|hr|hours)?$", low):
            numeric_or_unit += 1
        if low in {"yes", "no", "true", "false", "0", "1"}:
            bool_like += 1
        if any(ch in raw for ch in [",", ";", "|", "/"]):
            list_like += 1
        if re.search(r"\d+\s*[-to]{1,3}\s*\d+", low):
            range_like += 1
        if low.startswith("http://") or low.startswith("https://"):
            url_like += 1
        if re.search(r"\d{4}[-/]\d{1,2}([-/]\d{1,2})?$", low):
            date_like += 1
    total = max(len(values), 1)
    numeric_ratio = numeric_or_unit / total
    bool_ratio = bool_like / total
    list_ratio = list_like / total
    range_ratio = range_like / total
    url_ratio = url_like / total
    date_ratio = date_like / total

    inferred_type = "string"
    inferred_shape = "scalar"
    if bool_ratio >= 0.7:
        inferred_type = "boolean"
    elif url_ratio >= 0.5 or key_token.endswith("_link") or key_token.endswith("_url"):
        inferred_type = "url"
    elif date_ratio >= 0.5 or "date" in key_token:
        inferred_type = "date"
    elif range_ratio >= 0.35:
        inferred_type = "number"
        inferred_shape = "range"
    elif numeric_ratio >= 0.6:
        inferred_type = "number"
    elif list_ratio >= 0.5:
        inferred_shape = "list"

    if key_token in {"colors"} or "color" in key_token:
        inferred_type = "string"
        inferred_shape = "list"
    if key_token in {"polling_rate"}:
        inferred_type = "number"
        inferred_shape = "list"
    if "battery" in key_token and "hour" in key_token:
        inferred_type = "number"
        inferred_shape = "scalar"
    if key_token in {"cable_type", "sensor", "switch", "encoder", "material", "mcu"}:
        inferred_type = "string"
        inferred_shape = "scalar"

    unit_hint = infer_unit_hint(selected_key)
    parse_hint = infer_parse_template_hint(
        selected_key,
        inferred_type,
        inferred_shape,
        component_hint=any(tag in key_token for tag in ["sensor", "switch", "encoder", "material", "mcu"]),
    )
    value_form = "scalar"
    if inferred_shape == "list":
        value_form = "list"
    elif inferred_shape == "range":
        value_form = "range"
    return {
        "type": inferred_type,
        "shape": inferred_shape,
        "value_form": value_form,
        "unit": unit_hint,
        "parse_template": parse_hint,
        "numeric_ratio": numeric_ratio,
        "list_ratio": list_ratio,
        "bool_ratio": bool_ratio,
    }


def safe_json_parse(text: str, fallback):
    try:
        parsed = json.loads(str(text or "").strip() or "{}")
        if isinstance(fallback, dict):
            return parsed if isinstance(parsed, dict) else fallback
        if isinstance(fallback, list):
            return parsed if isinstance(parsed, list) else fallback
        return parsed
    except Exception:
        return fallback


def render_field_rules_studio(category: str, local_mode: bool):
    payload = load_studio_payload(category)
    paths = payload["paths"]
    compile_report = payload.get("compile_report", {}) if isinstance(payload.get("compile_report"), dict) else {}
    workbook_map = payload.get("workbook_map", {}) if isinstance(payload.get("workbook_map"), dict) else {}
    field_rules_draft = payload.get("field_rules_draft", {}) if isinstance(payload.get("field_rules_draft"), dict) else {}
    ui_field_catalog_draft = payload.get("ui_field_catalog_draft", {}) if isinstance(payload.get("ui_field_catalog_draft"), dict) else {}
    field_rows = field_rules_draft.get("fields", {}) if isinstance(field_rules_draft.get("fields"), dict) else {}
    generated_rules = payload.get("field_rules", {}) if isinstance(payload.get("field_rules"), dict) else {}
    known_values_payload = payload.get("known_values", {}) if isinstance(payload.get("known_values"), dict) else {}
    tooltip_source_path = str(payload.get("tooltip_source_path", "") or "")
    tooltip_entries = load_tooltip_bank_entries(tooltip_source_path)
    component_db_cache = {}

    def component_db_fallback(ctype: str):
        ckey = token(ctype)
        if ckey not in component_db_cache:
            component_db_cache[ckey] = component_db_snapshot(category, ctype)
        return component_db_cache.get(ckey, {"entity_count": 0, "sample_entities": []})

    def apply_tooltip_mapping(entries: dict):
        changed = False
        for key in list(field_rows.keys()):
            rule = field_rows.get(key, {})
            if not isinstance(rule, dict):
                continue
            ui = rule.get("ui", {}) if isinstance(rule.get("ui"), dict) else {}
            normalized_key = normalize_field_key(key)
            match = entries.get(normalized_key, {}) if isinstance(entries, dict) else {}
            if isinstance(match, dict) and match:
                tooltip_key = normalized_key
                tooltip_source = str(match.get("source", "") or "") or None
                if ui.get("tooltip_key") != tooltip_key:
                    ui["tooltip_key"] = tooltip_key
                    changed = True
                if ui.get("tooltip_source") != tooltip_source:
                    ui["tooltip_source"] = tooltip_source
                    changed = True
                if "tooltip_md" not in ui:
                    ui["tooltip_md"] = ""
                    changed = True
            else:
                if ui.get("tooltip_key"):
                    ui["tooltip_key"] = None
                    changed = True
                if ui.get("tooltip_source"):
                    ui["tooltip_source"] = None
                    changed = True
                if ui.get("tooltip_md", "") != "":
                    ui["tooltip_md"] = ""
                    changed = True
                if "tooltip_md" not in ui:
                    ui["tooltip_md"] = ""
                    changed = True
            rule["ui"] = ui
            field_rows[key] = rule
            upsert_ui_catalog_row(ui_field_catalog_draft, key, ui, rule)
        if changed:
            field_rules_draft["fields"] = field_rows
        return changed

    if not field_rows and isinstance(generated_rules.get("fields"), dict):
        field_rows = deepcopy(generated_rules.get("fields", {}))
        field_rules_draft = {**field_rules_draft, "category": category, "fields": field_rows}
        save_studio_drafts(category, field_rules_draft, ui_field_catalog_draft)

    compile_errors = compile_report.get("errors", []) if isinstance(compile_report.get("errors"), list) else []
    compile_warnings = compile_report.get("warnings", []) if isinstance(compile_report.get("warnings"), list) else []

    keys = sorted(field_rows.keys()) if isinstance(field_rows, dict) else []
    selected_key_state = f"studio_selected_key_{category}"
    if keys and st.session_state.get(selected_key_state) not in keys:
        st.session_state[selected_key_state] = keys[0]

    autosave_key = f"studio_autosave_{category}"
    if autosave_key not in st.session_state:
        st.session_state[autosave_key] = True

    st.subheader("Field Rules Studio")
    st.caption("Standalone contract authoring: mapping, key editor, workbook context, and deterministic compile.")

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Category", category)
    m2.metric("Contract Keys", len(keys))
    m3.metric("Compile Errors", len(compile_errors))
    m4.metric("Compile Warnings", len(compile_warnings))

    c0, c1, c2, c3 = st.columns([1.4, 1.1, 1.4, 1])
    c0.toggle("Auto Save on Change", key=autosave_key, help="When enabled, edits save immediately to draft files.")
    if c1.button("Save Draft", use_container_width=True, help="Save `_control_plane/field_rules_draft.json` and `_control_plane/ui_field_catalog_draft.json`."):
        save_studio_drafts(category, field_rules_draft, ui_field_catalog_draft)
        st.success("Draft files saved.")
    if c2.button("Compile & Generate Artifacts", use_container_width=True, help="Runs `category-compile` to write `_generated/*` from workbook + drafts + tooltip bank."):
        args = ["category-compile", "--category", category]
        if local_mode:
            args.append("--local")
        start_cli(args)
        st.info("Compile started. Watch Live Runtime / Process Output for progress.")
    if c3.button("Refresh", use_container_width=True):
        st.rerun()

    tab_map, tab_nav, tab_contract, tab_context, tab_reports = st.tabs(
        ["1) Mapping Studio", "2) Key Navigator", "3) Open Field Contract", "4) Workbook Context", "5) Compile & Reports"]
    )

    with tab_map:
        st.caption("Set workbook key source and sampling in plain-language controls.")
        mapped_sheet_names = collect_sheet_names(workbook_map)
        key_map = workbook_map.get("key_list", {}) if isinstance(workbook_map.get("key_list"), dict) else {}
        product_map = workbook_map.get("product_table", {}) if isinstance(workbook_map.get("product_table"), dict) else {}
        st.text_input("Workbook File", value=str(workbook_map.get("workbook_path", "") or ""), key=f"studio_workbook_path_{category}", help="Workbook path used by compile.")
        workbook_path_for_map = str(st.session_state.get(f"studio_workbook_path_{category}", workbook_map.get("workbook_path", "") or "")).strip()
        discovered_sheet_names = workbook_sheet_names(workbook_path_for_map)
        sheet_names = sorted(set(mapped_sheet_names + discovered_sheet_names)) or ["Sheet1"]
        k1, k2, k3, k4 = st.columns(4)
        k1.selectbox("Key Sheet", sheet_names, index=sheet_names.index(str(key_map.get("sheet", "") or sheet_names[0])) if str(key_map.get("sheet", "") or "") in sheet_names else 0, key=f"studio_key_sheet_{category}", help="Sheet containing source key list.")
        k2.text_input("Key Column", value=str(key_map.get("column", "B") or "B"), key=f"studio_key_col_{category}", help="Column containing key names.")
        k3.number_input("First Key Row", min_value=1, max_value=100000, value=int(key_map.get("row_start", 9) or 9), step=1, key=f"studio_key_row_start_{category}", help="First row for key extraction.")
        k4.number_input("Last Key Row", min_value=0, max_value=100000, value=int(key_map.get("row_end", 83) or 83), step=1, key=f"studio_key_row_end_{category}", help="0 means auto-detect until blank.")
        p1, p2, p3, p4 = st.columns(4)
        p1.selectbox("Sampling Sheet", sheet_names, index=sheet_names.index(str(product_map.get("sheet", "") or sheet_names[0])) if str(product_map.get("sheet", "") or "") in sheet_names else 0, key=f"studio_prod_sheet_{category}", help="Sheet used for sample values.")
        p2.selectbox("Layout", ["matrix", "row_table", "none"], index=["matrix", "row_table", "none"].index(str(product_map.get("layout", "matrix")) if str(product_map.get("layout", "matrix")) in ["matrix", "row_table", "none"] else "matrix"), key=f"studio_prod_layout_{category}", help="Matrix = keys down rows, products across columns.")
        p3.text_input("Value Start Column", value=str(product_map.get("value_col_start", "C") or "C"), key=f"studio_value_col_start_{category}", help="First product/value column.")
        p4.caption("Sampling mode: **All columns** from Value Start Column to last populated product column.")

        st.markdown("#### Tooltip Source")
        tooltip_cfg = workbook_map.get("tooltip_source", {}) if isinstance(workbook_map.get("tooltip_source"), dict) else {}
        default_tooltip_path = str(st.session_state.get(f"studio_tooltip_source_path_{category}", tooltip_cfg.get("path", tooltip_source_path) or tooltip_source_path or "")).strip()
        t1, t2, t3, t4 = st.columns([2.4, 1, 1, 1.2])
        tooltip_path_value = t1.text_input(
            "Tooltip Bank File (JS/JSON/MD)",
            value=default_tooltip_path,
            key=f"studio_tooltip_source_path_{category}",
            help="Path to tooltip bank file. Relative paths resolve from repo root.",
        )
        tooltip_entries_live = load_tooltip_bank_entries(tooltip_path_value)
        coverage_keys = [key for key in keys if normalize_field_key(key) in tooltip_entries_live]
        coverage_pct = (len(coverage_keys) / max(1, len(keys))) * 100.0
        t2.metric("Bank Keys", len(tooltip_entries_live))
        t3.metric("Coverage", f"{coverage_pct:.0f}%")
        if t4.button("Apply Tooltip Mapping", use_container_width=True, help="Bind tooltip_key for matched fields; set empty tooltip_md for missing keys."):
            changed = apply_tooltip_mapping(tooltip_entries_live)
            if changed:
                save_studio_drafts(category, field_rules_draft, ui_field_catalog_draft)
                st.success("Tooltip mapping applied to draft fields.")
                st.rerun()
            st.info("Tooltip mapping already up to date.")
        if tooltip_path_value and not tooltip_entries_live:
            st.warning("No tooltip entries parsed from the selected file.")
        elif tooltip_entries_live:
            st.caption("Parsed tooltip keys preview: " + ", ".join(sorted(list(tooltip_entries_live.keys()))[:12]))

        st.markdown("#### Component Source Mapping")
        st.caption("Universal component source mapping: required name/model column + row settings, optional role columns.")
        workbook_path_value = str(st.session_state.get(f"studio_workbook_path_{category}", workbook_map.get("workbook_path", "") or "")).strip()
        component_rows = normalize_component_rows(workbook_map)
        source_count_key = f"studio_component_source_count_{category}"
        if source_count_key not in st.session_state:
            st.session_state[source_count_key] = len(component_rows)
        s_c1, s_c2, s_c3 = st.columns([1, 1, 2.4])
        if s_c1.button("Add Source", use_container_width=True):
            st.session_state[source_count_key] = int(st.session_state.get(source_count_key, len(component_rows)) or len(component_rows)) + 1
            st.rerun()
        if s_c2.button("Remove Last", use_container_width=True, disabled=int(st.session_state.get(source_count_key, len(component_rows)) or len(component_rows)) <= 1):
            st.session_state[source_count_key] = max(1, int(st.session_state.get(source_count_key, len(component_rows)) or len(component_rows)) - 1)
            st.rerun()
        s_c3.caption("Add/remove component sources to map any category-specific component sheets.")
        target_count = max(1, int(st.session_state.get(source_count_key, len(component_rows)) or len(component_rows)))
        if len(component_rows) < target_count:
            for _ in range(target_count - len(component_rows)):
                component_rows.append(
                    {
                        "sheet": sheet_names[0] if sheet_names else "",
                        "component_type": "",
                        "header_row": 1,
                        "first_data_row": 2,
                        "canonical_name_column": "A",
                        "brand_column": "",
                        "alias_columns": [],
                        "link_columns": [],
                        "property_columns": [],
                        "auto_derive_aliases": True,
                        "stop_after_blank_names": 10,
                    }
                )
        component_rows = component_rows[:target_count]
        component_rows_out = []
        mapping_errors = []

        for idx, row in enumerate(component_rows):
            source_title = str(row.get("component_type", "") or f"source_{idx + 1}")
            st.markdown(f"##### `{source_title}`")
            m_c1, m_c2, m_c3, m_c4 = st.columns(4)
            sheet_default = str(row.get("sheet", "") or (sheet_names[idx] if idx < len(sheet_names) else sheet_names[0]))
            if sheet_default not in sheet_names:
                sheet_default = sheet_names[0]
            comp_type = m_c1.text_input(
                "Component Type",
                value=str(row.get("component_type", "") or ""),
                key=f"studio_comp_type_{category}_{idx}",
                help="Type token used for component_db output.",
            )
            comp_sheet = m_c2.selectbox(
                "Sheet",
                sheet_names,
                index=sheet_names.index(sheet_default),
                key=f"studio_comp_sheet_{category}_{idx}",
                help="Workbook sheet for this component type.",
            )
            header_row = m_c3.number_input(
                "Header Row",
                min_value=1,
                max_value=100000,
                value=int(row.get("header_row", 1) or 1),
                step=1,
                key=f"studio_comp_header_row_{category}_{idx}",
            )
            first_data_row = m_c4.number_input(
                "First Data Row",
                min_value=1,
                max_value=100000,
                value=int(row.get("first_data_row", 2) or 2),
                step=1,
                key=f"studio_comp_first_data_row_{category}_{idx}",
            )
            m_d1, m_d2, m_d3, m_d4 = st.columns(4)
            stop_after_blank_names = m_d1.number_input(
                "Stop After Blank Names",
                min_value=1,
                max_value=200,
                value=int(row.get("stop_after_blank_names", 10) or 10),
                step=1,
                key=f"studio_comp_stop_blank_{category}_{idx}",
            )
            selected_name_col_key = f"studio_comp_name_col_{category}_{idx}"
            name_input_key = f"studio_comp_name_col_input_{category}_{idx}"
            brand_input_key = f"studio_comp_brand_col_{category}_{idx}"
            alias_input_key = f"studio_comp_alias_cols_{category}_{idx}"
            link_input_key = f"studio_comp_link_cols_{category}_{idx}"
            property_input_key = f"studio_comp_prop_cols_{category}_{idx}"
            picker_target_key = f"studio_comp_picker_target_{category}_{idx}"
            active_roles_key = f"studio_comp_active_roles_{category}_{idx}"
            role_picker_key = f"studio_comp_role_picker_{category}_{idx}"
            add_role_key = f"studio_comp_add_role_{category}_{idx}"
            auto_alias_key = f"studio_comp_auto_alias_{category}_{idx}"
            clear_optional_key = f"studio_comp_clear_optional_{category}_{idx}"

            selected_name_col = normalize_col(
                st.session_state.get(selected_name_col_key, row.get("canonical_name_column", "A")),
                "A",
            )
            if selected_name_col_key not in st.session_state:
                st.session_state[selected_name_col_key] = selected_name_col
            if name_input_key not in st.session_state:
                st.session_state[name_input_key] = selected_name_col
            if brand_input_key not in st.session_state:
                st.session_state[brand_input_key] = normalize_col(row.get("brand_column", ""), "")
            if alias_input_key not in st.session_state:
                st.session_state[alias_input_key] = csv_join(stable_sort_columns(row.get("alias_columns", [])))
            if link_input_key not in st.session_state:
                st.session_state[link_input_key] = csv_join(stable_sort_columns(row.get("link_columns", [])))
            if property_input_key not in st.session_state:
                st.session_state[property_input_key] = csv_join(stable_sort_columns(row.get("property_columns", [])))
            if active_roles_key not in st.session_state:
                role_tokens = []
                if normalize_col(row.get("brand_column", ""), ""):
                    role_tokens.append("brand")
                if stable_sort_columns(row.get("link_columns", [])):
                    role_tokens.append("links")
                if stable_sort_columns(row.get("property_columns", [])):
                    role_tokens.append("properties")
                st.session_state[active_roles_key] = role_tokens
            if auto_alias_key not in st.session_state:
                st.session_state[auto_alias_key] = bool(row.get("auto_derive_aliases", True))
            role_defs = [
                {"id": "brand", "label": "Brand", "input_key": brand_input_key},
                {"id": "links", "label": "Links", "input_key": link_input_key},
                {"id": "properties", "label": "Properties", "input_key": property_input_key},
            ]
            role_def_by_id = {row["id"]: row for row in role_defs}
            known_role_ids = [row["id"] for row in role_defs]
            active_roles = [role for role in st.session_state.get(active_roles_key, []) if role in known_role_ids]
            active_role_set = set(active_roles)
            st.session_state[active_roles_key] = active_roles

            addable_roles = [role for role in role_defs if role["id"] not in active_role_set]
            role_mgr_a, role_mgr_b, role_mgr_c = st.columns([1.4, 1.3, 1.3])
            role_mgr_a.caption("Aliases role is always available.")
            if addable_roles:
                add_labels = [role["label"] for role in addable_roles]
                if role_picker_key not in st.session_state or st.session_state.get(role_picker_key) not in add_labels:
                    st.session_state[role_picker_key] = add_labels[0]
                selected_role_label = role_mgr_b.selectbox(
                    "Add Optional Role",
                    add_labels,
                    key=role_picker_key,
                    label_visibility="collapsed",
                )
                if role_mgr_c.button("Add Role", key=add_role_key, use_container_width=True):
                    selected_role = next((role for role in addable_roles if role["label"] == selected_role_label), None)
                    if selected_role:
                        active_role_set.add(selected_role["id"])
                        st.session_state[active_roles_key] = sorted(active_role_set)
                        st.rerun()
            else:
                role_mgr_b.caption("All optional roles added.")
                role_mgr_c.write("")

            if active_roles:
                st.caption("Active optional roles")
                remove_cols = st.columns(max(1, min(4, len(active_roles))))
                for ridx, role_id in enumerate(active_roles):
                    role_def = role_def_by_id.get(role_id, {})
                    label = str(role_def.get("label", role_id))
                    remove_key = f"studio_comp_remove_role_{category}_{idx}_{role_id}"
                    if remove_cols[ridx % len(remove_cols)].button(f"Remove {label}", key=remove_key, use_container_width=True):
                        active_role_set.discard(role_id)
                        st.session_state[active_roles_key] = sorted(active_role_set)
                        input_key = str(role_def.get("input_key", "") or "")
                        if input_key:
                            st.session_state[input_key] = ""
                        st.rerun()

            if st.button("Clear Optional Roles", key=clear_optional_key, use_container_width=True):
                st.session_state[active_roles_key] = []
                st.session_state[brand_input_key] = ""
                st.session_state[link_input_key] = ""
                st.session_state[property_input_key] = ""
                st.rerun()

            brand_enabled = "brand" in active_role_set
            link_enabled = "links" in active_role_set
            property_enabled = "properties" in active_role_set

            auto_derive_aliases = st.toggle(
                "Auto-Derive Aliases",
                key=auto_alias_key,
                help="If alias columns are empty, generate safe aliases from punctuation/spacing variants of Name/Model.",
            )
            st.caption(
                "Role-based mapping: `Name/Model` is required. Optional roles can stay off. "
                "Example: material often uses only `Name/Model`, optionally `Aliases`."
            )

            if not brand_enabled:
                st.session_state[brand_input_key] = ""
            if not link_enabled:
                st.session_state[link_input_key] = ""
            if not property_enabled:
                st.session_state[property_input_key] = ""

            preview = workbook_component_preview(workbook_path_value, comp_sheet, int(header_row), int(first_data_row))

            name_preview_values = []
            if "error" in preview:
                st.error(str(preview.get("error")))
                mapping_errors.append(f"{source_title}: {preview.get('error')}")
            else:
                preview_cols = preview.get("columns", []) if isinstance(preview.get("columns"), list) else []
                if preview_cols:
                    st.caption("Column Picker (click a header to map the selected role)")
                    picker_targets = ["Name/Model", "Aliases"]
                    if brand_enabled:
                        picker_targets.append("Brand")
                    if link_enabled:
                        picker_targets.append("Links")
                    if property_enabled:
                        picker_targets.append("Properties")
                    if picker_target_key not in st.session_state or st.session_state.get(picker_target_key) not in picker_targets:
                        st.session_state[picker_target_key] = picker_targets[0]
                    picker_target = st.selectbox(
                        "Mapping Target",
                        picker_targets,
                        key=picker_target_key,
                        help="Step 1: choose target role. Step 2: click a column header to assign/toggle it.",
                    )
                    btn_cols = st.columns(5)
                    for col_idx, col_meta in enumerate(preview_cols):
                        col_label = str(col_meta.get("col", "") or "")
                        header_label = str(col_meta.get("header", "") or "")
                        text = f"{col_label}: {header_label or '(blank)'}"
                        if btn_cols[col_idx % 5].button(text, key=f"studio_comp_pick_name_col_{category}_{idx}_{col_label}", use_container_width=True):
                            if picker_target == "Name/Model":
                                st.session_state[selected_name_col_key] = col_label
                                st.session_state[name_input_key] = col_label
                            elif picker_target == "Brand":
                                current = normalize_col(st.session_state.get(brand_input_key, ""), "")
                                st.session_state[brand_input_key] = "" if current == col_label else col_label
                            elif picker_target == "Aliases":
                                st.session_state[alias_input_key] = _toggle_csv_column(str(st.session_state.get(alias_input_key, "") or ""), col_label)
                            elif picker_target == "Links":
                                st.session_state[link_input_key] = _toggle_csv_column(str(st.session_state.get(link_input_key, "") or ""), col_label)
                            else:
                                st.session_state[property_input_key] = _toggle_csv_column(str(st.session_state.get(property_input_key, "") or ""), col_label)
                            st.rerun()
                    selected_name_col = normalize_col(st.session_state.get(selected_name_col_key, selected_name_col), selected_name_col)
                selected_brand_col_hint = normalize_col(st.session_state.get(brand_input_key, ""), "")
                selected_alias_cols_hint = stable_sort_columns(csv_tokens(str(st.session_state.get(alias_input_key, "") or "")))
                selected_link_cols_hint = stable_sort_columns(csv_tokens(str(st.session_state.get(link_input_key, "") or "")))
                selected_prop_cols_hint = stable_sort_columns(csv_tokens(str(st.session_state.get(property_input_key, "") or "")))
                st.caption(
                    "Selected mappings: "
                    f"name `{selected_name_col or '?'}` | "
                    f"brand `{selected_brand_col_hint or '-'}` | "
                    f"aliases `{csv_join(selected_alias_cols_hint) or '-'}` | "
                    f"links `{csv_join(selected_link_cols_hint) or '-'}` | "
                    f"properties `{csv_join(selected_prop_cols_hint) or '-'}`"
                )
                st.dataframe(preview.get("preview_rows", []), hide_index=True, use_container_width=True, height=200)

            manual_name_col = st.text_input(
                "Name/Model Column (required)",
                key=name_input_key,
                help="Single column letter. You can type it directly or pick via header click.",
            )
            selected_name_col = normalize_col(manual_name_col, "A")
            st.session_state[selected_name_col_key] = selected_name_col
            brand_column = m_d2.text_input(
                "Brand Column (optional)",
                key=brand_input_key,
                help="Optional column letter.",
                disabled=not brand_enabled,
            )
            alias_columns = m_d3.text_input(
                "Alias Columns (csv)",
                key=alias_input_key,
                help="Alias role is always available; leave blank if not needed.",
            )
            if link_enabled:
                link_columns = m_d4.text_input(
                    "Link Columns (optional csv)",
                    key=link_input_key,
                )
            else:
                m_d4.caption("Links role not added.")
                link_columns = ""
            if property_enabled:
                property_columns = st.text_input(
                    "Property Columns (optional csv)",
                    key=property_input_key,
                )
            else:
                st.caption("Properties role not added.")
                property_columns = ""

            selected_brand_col = normalize_col(brand_column if brand_enabled else "", "")
            selected_alias_cols = stable_sort_columns(csv_tokens(alias_columns))
            selected_link_cols = stable_sort_columns(csv_tokens(link_columns if link_enabled else ""))
            selected_property_cols = stable_sort_columns(csv_tokens(property_columns if property_enabled else ""))

            entity_preview = workbook_component_entities_preview(
                workbook_path=workbook_path_value,
                sheet_name=comp_sheet,
                header_row=int(header_row),
                first_data_row=int(first_data_row),
                canonical_name_column=selected_name_col,
                brand_column=selected_brand_col,
                alias_columns_csv=csv_join(selected_alias_cols),
                link_columns_csv=csv_join(selected_link_cols),
                auto_derive_aliases=bool(auto_derive_aliases),
                stop_after_blank_names=int(stop_after_blank_names),
            )
            if "error" in entity_preview:
                st.error(str(entity_preview.get("error")))
                mapping_errors.append(f"{source_title}: {entity_preview.get('error')}")
            else:
                preview_rows = entity_preview.get("preview_rows", []) if isinstance(entity_preview.get("preview_rows"), list) else []
                name_preview_values = entity_preview.get("first_20_names", []) if isinstance(entity_preview.get("first_20_names"), list) else []
                st.caption("Live Preview (first 20 rows): {name, brand, aliases[], links[]}")
                st.dataframe(preview_rows, hide_index=True, use_container_width=True, height=260)
                sample_names = entity_preview.get("sample_names", []) if isinstance(entity_preview.get("sample_names"), list) else []
                st.caption(
                    f"Entity count: {int(entity_preview.get('entity_count', 0) or 0)} | "
                    f"Samples: {', '.join(sample_names[:10]) or 'n/a'}"
                )
                preview_errors = component_preview_errors(name_preview_values)
                if float(entity_preview.get("numeric_only_ratio", 0.0) or 0.0) > 0.10:
                    preview_errors.append("More than 10% of component names are numeric-only. Choose the Name/Model column, not the ID column.")
                if bool(entity_preview.get("first_20_all_numeric", False)):
                    preview_errors.append("The first 20 component names are numeric-only. Choose the Name/Model column.")
                for err in ordered_unique_text(preview_errors):
                    st.error(err)
                    mapping_errors.append(f"{source_title}: {err}")
                if not preview_errors:
                    ratio = float(entity_preview.get("numeric_only_ratio", 0.0) or 0.0)
                    st.success(
                        f"Preview looks valid ({int(entity_preview.get('entity_count', 0) or 0)} entities, "
                        f"numeric-only ratio {ratio:.0%})."
                    )

            component_rows_out.append(
                {
                    "sheet": str(comp_sheet or "").strip(),
                    "component_type": str(comp_type or "").strip(),
                    "header_row": int(header_row),
                    "first_data_row": int(first_data_row),
                    "canonical_name_column": normalize_col(st.session_state.get(selected_name_col_key, selected_name_col), "A"),
                    "brand_column": selected_brand_col,
                    "alias_columns": selected_alias_cols,
                    "link_columns": selected_link_cols,
                    "property_columns": selected_property_cols,
                    "auto_derive_aliases": bool(auto_derive_aliases),
                    "stop_after_blank_names": int(stop_after_blank_names),
                }
            )

        mapping_can_save = len(mapping_errors) == 0
        if not mapping_can_save:
            st.warning("Fix component source mapping preview errors before saving.")
            st.dataframe([{"issue": issue} for issue in mapping_errors], hide_index=True, use_container_width=True, height=180)

        if st.button(
            "Save Mapping",
            use_container_width=True,
            disabled=not mapping_can_save,
            help="Save workbook key/product mapping and component source mapping to `_control_plane/workbook_map.json`.",
        ):
            tooltip_path_to_store = str(st.session_state.get(f"studio_tooltip_source_path_{category}", tooltip_path_value or "") or "").strip()
            tooltip_format = Path(tooltip_path_to_store).suffix.lower().lstrip(".") if tooltip_path_to_store else "auto"
            workbook_map = {
                **workbook_map,
                "workbook_path": workbook_path_value,
                "tooltip_source": {
                    "path": tooltip_path_to_store,
                    "format": tooltip_format or "auto",
                },
                "key_list": {
                    **key_map,
                    "sheet": str(st.session_state.get(f"studio_key_sheet_{category}", key_map.get("sheet", "dataEntry"))),
                    "column": str(st.session_state.get(f"studio_key_col_{category}", key_map.get("column", "B")) or "B"),
                    "row_start": int(st.session_state.get(f"studio_key_row_start_{category}", key_map.get("row_start", 9) or 9)),
                    "row_end": int(st.session_state.get(f"studio_key_row_end_{category}", key_map.get("row_end", 83) or 83)),
                    "source": "column_range",
                },
                "product_table": {
                    **product_map,
                    "sheet": str(st.session_state.get(f"studio_prod_sheet_{category}", product_map.get("sheet", "dataEntry"))),
                    "layout": str(st.session_state.get(f"studio_prod_layout_{category}", product_map.get("layout", "matrix"))),
                    "key_column": str(st.session_state.get(f"studio_key_col_{category}", key_map.get("column", "B")) or "B"),
                    "value_col_start": str(st.session_state.get(f"studio_value_col_start_{category}", product_map.get("value_col_start", "C") or "C")),
                    "sample_columns": 0,
                },
                "component_sheets": component_rows_out,
                "component_sources": [
                    {
                        "sheet": row.get("sheet"),
                        "type": row.get("component_type"),
                        "header_row": row.get("header_row"),
                        "first_data_row": row.get("first_data_row"),
                        "canonical_name_column": row.get("canonical_name_column"),
                        "brand_column": row.get("brand_column") or None,
                        "alias_columns": row.get("alias_columns", []),
                        "link_columns": row.get("link_columns", []),
                        "property_columns": row.get("property_columns", []),
                        "auto_derive_aliases": bool(row.get("auto_derive_aliases", True)),
                        "stop_after_blank_names": row.get("stop_after_blank_names", 10),
                    }
                    for row in component_rows_out
                    if str(row.get("sheet", "")).strip()
                ],
            }
            write_json(paths["workbook_map"], workbook_map)
            if tooltip_path_to_store and apply_tooltip_mapping(tooltip_entries_live):
                save_studio_drafts(category, field_rules_draft, ui_field_catalog_draft)
            st.success("Mapping saved.")

        st.markdown("#### Globals / Sources")
        workbook_tabs = generated_rules.get("workbook_tabs", {}) if isinstance(generated_rules.get("workbook_tabs"), dict) else {}
        enum_buckets = generated_rules.get("enum_buckets", {}) if isinstance(generated_rules.get("enum_buckets"), dict) else {}
        component_sources = generated_rules.get("component_db_sources", {}) if isinstance(generated_rules.get("component_db_sources"), dict) else {}
        parse_templates = generated_rules.get("parse_templates", {}) if isinstance(generated_rules.get("parse_templates"), dict) else {}

        with st.expander("Workbook Tabs", expanded=False):
            if workbook_tabs:
                rows = []
                for sheet, meta in workbook_tabs.items():
                    meta = meta if isinstance(meta, dict) else {}
                    rows.append(
                        {
                            "sheet": sheet,
                            "role": str(meta.get("role", "")),
                            "notes": str(meta.get("notes", "")),
                        }
                    )
                st.dataframe(rows, hide_index=True, use_container_width=True, height=220)
            else:
                st.info("Compile once to populate workbook tab role summary.")

        with st.expander("Enum Buckets", expanded=False):
            if enum_buckets:
                rows = []
                for bucket, meta in enum_buckets.items():
                    meta = meta if isinstance(meta, dict) else {}
                    excel_meta = meta.get("excel", {}) if isinstance(meta.get("excel"), dict) else {}
                    values = meta.get("values", []) if isinstance(meta.get("values"), list) else []
                    rows.append(
                        {
                            "bucket": bucket,
                            "sheet": str(excel_meta.get("sheet", "")),
                            "column": str(excel_meta.get("column", "")),
                            "value_count": len(values),
                            "preview": ", ".join([str(v) for v in values[:6]]),
                        }
                    )
                st.dataframe(rows, hide_index=True, use_container_width=True, height=240)
            else:
                st.info("No enum buckets generated yet.")

        with st.expander("Component Sources", expanded=False):
            if component_sources:
                rows = []
                for ctype, meta in component_sources.items():
                    meta = meta if isinstance(meta, dict) else {}
                    excel_meta = meta.get("excel", {}) if isinstance(meta.get("excel"), dict) else {}
                    if not excel_meta:
                        excel_meta = {
                            "sheet": meta.get("sheet", ""),
                            "canonical_name_column": meta.get("canonical_name_column", meta.get("name_column", "")),
                            "name_column": meta.get("name_column", ""),
                            "alias_columns": meta.get("alias_columns", []),
                        }
                    sample_entities = meta.get("sample_entities", []) if isinstance(meta.get("sample_entities"), list) else []
                    entity_count = int(meta.get("entity_count", 0) or 0)
                    fallback = component_db_fallback(str(ctype))
                    if not sample_entities:
                        sample_entities = fallback.get("sample_entities", [])
                    if entity_count <= 0:
                        entity_count = int(fallback.get("entity_count", 0) or 0)
                    if entity_count <= 0:
                        entity_count = len(sample_entities)
                    rows.append(
                        {
                            "type": ctype,
                            "sheet": str(excel_meta.get("sheet", "")),
                            "name_col": str(excel_meta.get("canonical_name_column", "") or excel_meta.get("name_column", "")),
                            "aliases": ", ".join(excel_meta.get("alias_columns", []) if isinstance(excel_meta.get("alias_columns"), list) else []),
                            "entity_count": entity_count,
                        }
                    )
                st.dataframe(rows, hide_index=True, use_container_width=True, height=220)
            else:
                st.info("No component sources generated yet.")

        with st.expander("Parse Templates", expanded=False):
            if parse_templates:
                rows = []
                for name, meta in parse_templates.items():
                    meta = meta if isinstance(meta, dict) else {}
                    rows.append(
                        {
                            "template": name,
                            "description": str(meta.get("description", "")),
                        }
                    )
                st.dataframe(rows, hide_index=True, use_container_width=True, height=240)
            else:
                st.info("Parse template catalog will appear after compile.")

    if not isinstance(field_rows, dict) or not field_rows:
        with tab_contract:
            st.info("No draft keys found yet. Run `Compile & Generate Artifacts` to bootstrap drafts from workbook + tooltip bank.")
        return

    enum_bucket_meta = generated_rules.get("enum_buckets", {}) if isinstance(generated_rules.get("enum_buckets"), dict) else {}
    enum_fields = sorted([str(k) for k in enum_bucket_meta.keys() if str(k).strip()])
    if not enum_fields and isinstance(known_values_payload.get("fields"), dict):
        enum_fields = sorted([str(k) for k in known_values_payload.get("fields", {}).keys() if str(k).strip()])

    component_source_meta = generated_rules.get("component_db_sources", {}) if isinstance(generated_rules.get("component_db_sources"), dict) else {}
    component_types = sorted([str(k) for k in component_source_meta.keys() if str(k).strip()])
    if not component_types:
        for row in workbook_map.get("component_sheets", []) if isinstance(workbook_map.get("component_sheets"), list) else []:
            if isinstance(row, dict) and str(row.get("component_type", "")).strip():
                component_types.append(str(row.get("component_type")).strip())
        component_types = sorted(set(component_types))

    with tab_nav:
        nav_rows = []
        for key in keys:
            rule = field_rows.get(key, {})
            priority = rule_priority(rule)
            contract = rule_contract(rule)
            errs = key_contract_errors(rule)
            nav_rows.append(
                {
                    "open": key == st.session_state.get(selected_key_state, key),
                    "key": key,
                    "status": "ERR" if errs else "OK",
                    "required_level": str(priority.get("required_level", "optional") or "optional"),
                    "type": str(contract.get("type", "")),
                    "shape": str(contract.get("shape", "")),
                    "issue": errs[0] if errs else "Field override appears complete.",
                }
            )
        search = st.text_input("Search", value="", key=f"studio_nav_search_{category}", help="Filter by key/status/type.").strip().lower()
        only_errors = st.toggle("Errors Only", value=False, key=f"studio_nav_errors_{category}")
        filtered = []
        for row in nav_rows:
            blob = f"{row['key']} {row['status']} {row['required_level']} {row['type']} {row['shape']}".lower()
            if search and search not in blob:
                continue
            if only_errors and row["status"] != "ERR":
                continue
            filtered.append(row)
        edited = st.data_editor(
            filtered if filtered else nav_rows,
            hide_index=True,
            use_container_width=True,
            disabled=["key", "status", "required_level", "type", "shape", "issue"],
            key=f"studio_key_table_{category}",
            column_config={
                "open": st.column_config.CheckboxColumn("Open", help="Select key to edit."),
                "issue": st.column_config.TextColumn("Key Status", width="large"),
            },
            height=440,
        )
        chosen = [row.get("key") for row in edited if bool(row.get("open")) and str(row.get("key", "")).strip()]
        if chosen:
            next_key = str(chosen[-1])
            if next_key in keys and next_key != st.session_state.get(selected_key_state):
                st.session_state[selected_key_state] = next_key
                st.rerun()

    selected_key = st.session_state.get(selected_key_state, keys[0])
    rule = deepcopy(field_rows.get(selected_key, {}))
    ui_row = deepcopy(rule.get("ui", {})) if isinstance(rule.get("ui"), dict) else {}
    priority = rule_priority(rule)
    contract = rule_contract(rule)
    parse_block = rule_parse(rule)
    enum_block = rule_enum(rule)

    with tab_contract:
        key_choice = st.selectbox(
            "Active Key",
            keys,
            index=keys.index(selected_key) if selected_key in keys else 0,
            key=f"studio_active_key_select_{category}",
            help="Choose any key directly for editing.",
        )
        if key_choice != selected_key:
            st.session_state[selected_key_state] = key_choice
            st.rerun()
        selected_key = st.session_state.get(selected_key_state, key_choice)
        rule = deepcopy(field_rows.get(selected_key, {}))
        ui_row = deepcopy(rule.get("ui", {})) if isinstance(rule.get("ui"), dict) else {}
        priority = rule_priority(rule)
        contract = rule_contract(rule)
        parse_block = rule_parse(rule)
        enum_block = rule_enum(rule)
        st.markdown(f"### Open Field Contract: `{selected_key}`")
        st.caption("Use dropdowns and templates to define this key. Hover each field label for guidance.")
        b1, b2, b3, b4 = st.columns(4)
        label = b1.text_input("Label", value=str(ui_row.get("label", selected_key) or selected_key), help="Human-readable key name.")
        group = b2.text_input("Group", value=str(ui_row.get("group", "General") or "General"), help="UI/catalog grouping.")
        order = b3.number_input("Order", min_value=1, max_value=5000, value=int(ui_row.get("order", 1) or 1), step=1, help="Stable ordering number.")
        canonical_key = b4.text_input("Canonical Key", value=str(rule.get("canonical_key", "") or ""), help="Optional canonical output key for migrations.")
        aliases = st.text_input("Aliases (comma separated)", value=csv_join(rule.get("aliases", [])), help="Synonyms used for extraction mapping.")
        tip_keys = [""] + sorted(tooltip_entries.keys())
        current_tip = str(ui_row.get("tooltip_key", "") or "")
        if current_tip not in tip_keys:
            tip_keys.append(current_tip)
        t1, t2 = st.columns(2)
        tooltip_key = t1.selectbox("Tooltip Key", tip_keys, index=tip_keys.index(current_tip) if current_tip in tip_keys else 0, help="Choose tooltip from tooltip bank if available.")
        tooltip_md = t2.text_area("Tooltip Markdown", value=str(ui_row.get("tooltip_md", "") or ""), height=140, help="Fallback tooltip content (must exist even if empty).")
        if tooltip_key and tooltip_key in tooltip_entries:
            tip = tooltip_entries[tooltip_key]
            st.markdown(f'<div style="max-height:220px;overflow:auto;border:1px solid #333;padding:10px;border-radius:8px;">{tip.get("html") or tip.get("plain","")}</div>', unsafe_allow_html=True)

        i1, i2, i3, i4 = st.columns(4)
        required_level = i1.selectbox("Required Level", ["identity", "required", "critical", "expected", "optional", "rare"], index=max(0, ["identity", "required", "critical", "expected", "optional", "rare"].index(str(priority.get("required_level", "optional")) if str(priority.get("required_level", "optional")) in ["identity", "required", "critical", "expected", "optional", "rare"] else "optional")), help="Priority and publish-gate severity.")
        availability = i2.selectbox("Availability", ["expected", "sometimes", "rare"], index=max(0, ["expected", "sometimes", "rare"].index(str(priority.get("availability", "sometimes")) if str(priority.get("availability", "sometimes")) in ["expected", "sometimes", "rare"] else "sometimes")), help="How often this field should be available publicly.")
        difficulty = i3.selectbox("Difficulty", ["easy", "medium", "hard"], index=max(0, ["easy", "medium", "hard"].index(str(priority.get("difficulty", "medium")) if str(priority.get("difficulty", "medium")) in ["easy", "medium", "hard"] else "medium")), help="Extraction complexity for this field.")
        effort = i4.slider("Effort", min_value=1, max_value=10, value=int(priority.get("effort", 5) or 5), help="Search/extraction effort budget.")

        c1, c2, c3, c4 = st.columns(4)
        ftype_options = ["string", "number", "integer", "boolean", "date", "url", "object"]
        shape_options = ["scalar", "list", "object", "range"]
        value_form_options = ["single", "set", "range", "mixed"]
        ftype = c1.selectbox("Type", ftype_options, index=max(0, ftype_options.index(str(contract.get("type", "string")) if str(contract.get("type", "string")) in ftype_options else "string")), help="Normalized primitive type.")
        shape = c2.selectbox("Shape", shape_options, index=max(0, shape_options.index(str(contract.get("shape", "scalar")) if str(contract.get("shape", "scalar")) in shape_options else "scalar")), help="Output shape contract.")
        value_form = c3.selectbox("Value Form", value_form_options, index=max(0, value_form_options.index(str(rule.get("value_form", "single")) if str(rule.get("value_form", "single")) in value_form_options else "single")), help="Scalar/list/range behavior.")
        unit = c4.text_input("Unit", value=str(contract.get("unit", "") or ""), help="Normalized unit (mm/g/hz/ms/dpi/h/none).")

        parse_template_options = [
            "text_field",
            "string",
            "enum_string",
            "boolean_yes_no_unk",
            "boolean_yes_no_unknown",
            "number_with_unit",
            "integer_with_unit",
            "list_of_tokens_delimited",
            "list_of_numbers_with_unit",
            "range_number",
            "mode_tagged_values",
            "mode_tagged_list",
            "url_field",
            "date_field",
            "component_reference",
            "latency_list_modes_ms",
            "list_numbers_or_ranges_with_unit",
        ]
        enum_policy_options = ["closed", "open_prefer_known", "open", "closed_with_curation"]
        p1, p2, p3, p4 = st.columns(4)
        parse_template = p1.selectbox(
            "Parse Template",
            parse_template_options,
            index=max(0, parse_template_options.index(str(parse_block.get("template", "text_field")) if str(parse_block.get("template", "text_field")) in parse_template_options else "text_field")),
            help="Deterministic parser template.",
        )
        enum_policy = p2.selectbox(
            "Enum Policy",
            enum_policy_options,
            index=max(0, enum_policy_options.index(str(enum_block.get("policy", "open_prefer_known")) if str(enum_block.get("policy", "open_prefer_known")) in enum_policy_options else "open_prefer_known")),
            help="Closed/open behavior for canonical value matching.",
        )
        enum_source_labels = {"none": "none"}
        enum_source_options = ["none"]
        for field in enum_fields:
            option = f"enum_buckets:{field}"
            meta = enum_bucket_meta.get(field, {}) if isinstance(enum_bucket_meta.get(field), dict) else {}
            excel_meta = meta.get("excel", {}) if isinstance(meta.get("excel"), dict) else {}
            count = len(meta.get("values", []) if isinstance(meta.get("values"), list) else [])
            sheet = str(excel_meta.get("sheet", "") or "")
            column = str(excel_meta.get("column", "") or "")
            enum_source_labels[option] = f"{option} ({sheet}:{column}, {count} values)" if sheet or column else f"{option} ({count} values)"
            enum_source_options.append(option)
        for ctype in component_types:
            option = f"component_db:{ctype}"
            meta = component_source_meta.get(ctype, {}) if isinstance(component_source_meta.get(ctype), dict) else {}
            excel_meta = meta.get("excel", {}) if isinstance(meta.get("excel"), dict) else {}
            if not excel_meta:
                excel_meta = {
                    "sheet": meta.get("sheet", ""),
                    "canonical_name_column": meta.get("canonical_name_column", meta.get("name_column", "")),
                }
            sheet = str(excel_meta.get("sheet", "") or "")
            count = int(meta.get("entity_count", 0) or 0)
            if count <= 0:
                count = int(component_db_fallback(ctype).get("entity_count", 0) or 0)
            enum_source_labels[option] = f"{option} ({sheet}, {count} entities)" if sheet else f"{option} ({count} entities)"
            enum_source_options.append(option)
        current_enum_source = normalize_enum_source_value(enum_block.get("source"))
        if current_enum_source not in enum_source_options:
            enum_source_options.append(current_enum_source)
            enum_source_labels[current_enum_source] = current_enum_source
        enum_source_value = p3.selectbox(
            "Enum Source",
            enum_source_options,
            index=enum_source_options.index(current_enum_source) if current_enum_source in enum_source_options else 0,
            format_func=lambda val: enum_source_labels.get(val, val),
            help="Bind to enum bucket or component source.",
        )
        min_refs = p4.selectbox("Min Evidence Refs", [1, 2], index=0 if int(rule_min_evidence(rule) or 1) == 1 else 1, help="Required snippet refs per accepted value.")
        q1, q2 = st.columns(2)
        strict_unit_required = q1.selectbox("Strict Unit Required", [True, False], index=0 if bool(parse_block.get("strict_unit_required", False)) else 1, help="Require explicit unit when conversions apply.")
        display_mode = q2.selectbox(
            "Display Mode",
            ["all", "range", "high", "low"],
            index=max(
                0,
                ["all", "range", "high", "low"].index(
                    str(ui_row.get("display_mode", "all"))
                    if str(ui_row.get("display_mode", "all")) in ["all", "range", "high", "low"]
                    else "all"
                ),
            ),
            help="Numeric display handling.",
        )
        r1, r2, r3 = st.columns(3)
        evidence_block = rule.get("evidence", {}) if isinstance(rule.get("evidence"), dict) else {}
        evidence_required = r1.toggle("Evidence Required", value=bool(evidence_block.get("required", rule.get("evidence_required", True))), help="Non-unk values require evidence.")
        publish_gate = r2.toggle("Publish Gate", value=bool(priority.get("publish_gate", False)), help="Field participates in publish gating.")
        block_publish_when_unk = r3.toggle("Block Publish When Unknown", value=bool(priority.get("block_publish_when_unk", False)), help="Unknown value blocks publish when enabled.")

        contract_rounding = contract.get("rounding", {}) if isinstance(contract.get("rounding"), dict) else {}
        contract_range = contract.get("range", {}) if isinstance(contract.get("range"), dict) else {}
        list_rules = contract.get("list_rules", {}) if isinstance(contract.get("list_rules"), dict) else (rule.get("list_rules", {}) if isinstance(rule.get("list_rules"), dict) else {})
        parse_rules = rule.get("parse_rules", {}) if isinstance(rule.get("parse_rules"), dict) else {}
        enum_match = enum_block.get("match", {}) if isinstance(enum_block.get("match"), dict) else {}
        new_value_policy_current = enum_block.get("new_value_policy", {}) if isinstance(enum_block.get("new_value_policy"), dict) else (rule.get("new_value_policy", {}) if isinstance(rule.get("new_value_policy"), dict) else {})
        component_block = rule.get("component", {}) if isinstance(rule.get("component"), dict) else {}
        selection_policy_block = rule.get("selection_policy", {}) if isinstance(rule.get("selection_policy"), dict) else {}
        search_hints_block = rule.get("search_hints", {}) if isinstance(rule.get("search_hints"), dict) else {}
        excel_hints_block = rule.get("excel_hints", {}) if isinstance(rule.get("excel_hints"), dict) else {}

        with st.expander("Advanced Contract Settings", expanded=False):
            a1, a2, a3, a4 = st.columns(4)
            round_decimals = a1.selectbox(
                "Rounding Decimals",
                [0, 1, 2, 3],
                index=max(0, [0, 1, 2, 3].index(int(contract_rounding.get("decimals", 0) or 0)) if int(contract_rounding.get("decimals", 0) or 0) in [0, 1, 2, 3] else 0),
                help="Decimal places for numeric normalization.",
            )
            round_mode = a2.selectbox(
                "Rounding Mode",
                ["nearest", "floor", "ceil"],
                index=max(0, ["nearest", "floor", "ceil"].index(str(contract_rounding.get("mode", "nearest")) if str(contract_rounding.get("mode", "nearest")) in ["nearest", "floor", "ceil"] else "nearest")),
                help="Numeric rounding policy.",
            )
            range_min = a3.text_input("Min Bound", value="" if contract_range.get("min") is None else str(contract_range.get("min")), help="Optional minimum numeric bound.")
            range_max = a4.text_input("Max Bound", value="" if contract_range.get("max") is None else str(contract_range.get("max")), help="Optional maximum numeric bound.")

            l1, l2, l3, l4 = st.columns(4)
            list_dedupe = l1.selectbox("List Dedupe", [True, False], index=0 if bool(list_rules.get("dedupe", True)) else 1, help="Deduplicate list outputs.")
            list_sort = l2.selectbox("List Sort", ["none", "asc", "desc"], index=max(0, ["none", "asc", "desc"].index(str(list_rules.get("sort", "none")) if str(list_rules.get("sort", "none")) in ["none", "asc", "desc"] else "none")), help="Sort order for list outputs.")
            list_min_items = l3.number_input("List Min Items", min_value=0, max_value=1000, value=int(list_rules.get("min_items", 0) or 0), step=1, help="Minimum list length.")
            list_max_items = l4.number_input("List Max Items", min_value=1, max_value=5000, value=int(list_rules.get("max_items", 100) or 100), step=1, help="Maximum list length.")

            object_schema_default = contract.get("object_schema") if isinstance(contract.get("object_schema"), dict) else (rule.get("object_schema") if isinstance(rule.get("object_schema"), dict) else {})
            object_schema_text = st.text_area(
                "Object Schema JSON",
                value=json.dumps(object_schema_default or {}, ensure_ascii=False, indent=2),
                height=140,
                help="Required when shape/object fields use structured values (e.g. latency lists).",
            )
            item_union_default = contract.get("item_union") if isinstance(contract.get("item_union"), list) else (rule.get("item_union") if isinstance(rule.get("item_union"), list) else [])
            item_union_text = st.text_area(
                "Item Union JSON",
                value=json.dumps(item_union_default or [], ensure_ascii=False, indent=2),
                height=120,
                help="Optional union contract for mixed list/range values.",
            )

            p_a1, p_a2, p_a3 = st.columns(3)
            unit_accepts_text = p_a1.text_input(
                "Unit Accepts (csv)",
                value=csv_join(parse_block.get("unit_accepts", parse_rules.get("unit_accepts", []))),
                help="Accepted input units before conversion (e.g. mm,cm,in,\").",
            )
            delimiters_text = p_a2.text_input(
                "Delimiters (csv)",
                value=csv_join(parse_block.get("delimiters", parse_rules.get("delimiters", []))),
                help="List separators for tokenized/list parsing.",
            )
            range_separators_text = p_a3.text_input(
                "Range Separators (csv)",
                value=csv_join(parse_block.get("range_separators", parse_rules.get("separators", []))),
                help="Range separators for range parsing (e.g. -,to).",
            )
            p_a4, p_a5 = st.columns(2)
            accepted_formats_text = p_a4.text_input(
                "Accepted Date Formats (csv)",
                value=csv_join(parse_block.get("accepted_formats", parse_rules.get("accepted_formats", []))),
                help="Date formats accepted by date parser.",
            )
            unit_conversions_text = p_a5.text_area(
                "Unit Conversions (k=v per line)",
                value=format_key_value_lines(parse_block.get("unit_conversions", parse_rules.get("unit_conversions", {}))),
                height=110,
                help="Example: in_to_mm=25.4",
            )
            token_map_text = st.text_area(
                "Token Map (k=v per line)",
                value=format_key_value_lines(parse_block.get("token_map", parse_rules.get("token_map", {}))),
                height=110,
                help="Optional token normalization map (e.g. grey=gray).",
            )
            pp1, pp2, pp3 = st.columns(3)
            allow_unitless = pp1.selectbox(
                "Allow Unitless",
                [True, False],
                index=0 if bool(parse_block.get("allow_unitless", parse_rules.get("allow_unitless", True))) else 1,
                help="Allow numbers without explicit units when parser supports it.",
            )
            allow_ranges = pp2.selectbox(
                "Allow Ranges",
                [True, False],
                index=0 if bool(parse_block.get("allow_ranges", parse_rules.get("allow_ranges", False))) else 1,
                help="Allow explicit range tokens (e.g. 1-3).",
            )
            accept_bare_mode = pp3.selectbox(
                "Accept Bare Numbers As Mode",
                [True, False],
                index=0 if bool(parse_block.get("accept_bare_numbers_as_mode", parse_rules.get("accept_bare_numbers_as_mode", False))) else 1,
                help="For mode-tagged parsers, allow bare numbers to imply current mode.",
            )
            mode_aliases_text = st.text_area(
                "Mode Aliases (k=v per line)",
                value=format_key_value_lines(parse_block.get("mode_aliases", parse_rules.get("mode_aliases", {}))),
                height=110,
                help="Optional mode alias normalization map.",
            )

            e_a1, e_a2, e_a3 = st.columns(3)
            enum_match_default = str(enum_match.get("strategy", rule.get("enum_match_strategy", "alias")) or "alias")
            if enum_match_default not in {"exact", "alias", "fuzzy"}:
                enum_match_default = "alias"
            enum_match_strategy = e_a1.selectbox(
                "Enum Match Strategy",
                ["exact", "alias", "fuzzy"],
                index=["exact", "alias", "fuzzy"].index(enum_match_default),
                help="Matching strategy for enum normalization.",
            )
            enum_fuzzy_threshold = e_a2.number_input(
                "Enum Fuzzy Threshold",
                min_value=0.0,
                max_value=1.0,
                value=float(enum_match.get("fuzzy_threshold", rule.get("enum_fuzzy_threshold", 0.92)) or 0.92),
                step=0.01,
                help="Only used when match strategy is fuzzy.",
            )
            new_value_target = e_a3.text_input(
                "New Value Suggestion Target",
                value=str(new_value_policy_current.get("suggestion_target", "_suggestions/enums.json") or "_suggestions/enums.json"),
                help="Where open-enum curation suggestions are written.",
            )
            e_b1, e_b2 = st.columns(2)
            nvp_accept = e_b1.selectbox("Accept New Enum Value If Evidence", [True, False], index=0 if bool(new_value_policy_current.get("accept_if_evidence", True)) else 1, help="Open enum behavior.")
            nvp_curation = e_b2.selectbox("Mark New Enum Value Needs Curation", [True, False], index=0 if bool(new_value_policy_current.get("mark_needs_curation", True)) else 1, help="Flag unknown enum tokens for review.")

            c_a1, c_a2, c_a3, c_a4 = st.columns(4)
            component_default = str(component_block.get("type", "") or "")
            if current_enum_source.startswith("component_db:") and not component_default:
                component_default = current_enum_source.split(":", 1)[1]
            component_type_option = c_a1.selectbox(
                "Component Type",
                ["none"] + component_types + ["custom"],
                index=max(0, (["none"] + component_types + ["custom"]).index(component_default if component_default in component_types else ("custom" if component_default else "none"))),
                help="Bind this key to a component DB type when applicable.",
            )
            component_custom_type = ""
            if component_type_option == "custom":
                component_custom_type = c_a2.text_input("Custom Component Type", value=component_default or "", help="Custom component type token.")
            component_require_identity = c_a3.selectbox("Require Component Identity Evidence", [True, False], index=0 if bool(component_block.get("require_identity_evidence", True)) else 1, help="Require explicit evidence for component identity.")
            component_allow_new = c_a4.selectbox("Allow New Components", [True, False], index=0 if bool(component_block.get("allow_new_components", True)) else 1, help="Allow unseen component names when evidence-backed.")

            ev_tier_pref_default = evidence_block.get("tier_preference", ["tier1", "tier2", "tier3"]) if isinstance(evidence_block.get("tier_preference", ["tier1", "tier2", "tier3"]), list) else ["tier1", "tier2", "tier3"]
            ev_conflict_default = str(evidence_block.get("conflict_policy", "resolve_by_tier_else_unknown") or "resolve_by_tier_else_unknown")
            ev1, ev2 = st.columns(2)
            evidence_tier_pref = ev1.multiselect(
                "Evidence Tier Preference",
                ["tier1", "tier2", "tier3"],
                default=[tier for tier in ev_tier_pref_default if tier in {"tier1", "tier2", "tier3"}] or ["tier1", "tier2", "tier3"],
                help="Source tier preference order for conflict resolution.",
            )
            conflict_policy = ev2.selectbox(
                "Conflict Policy",
                ["resolve_by_tier_else_unknown", "preserve_all_candidates"],
                index=max(0, ["resolve_by_tier_else_unknown", "preserve_all_candidates"].index(ev_conflict_default) if ev_conflict_default in ["resolve_by_tier_else_unknown", "preserve_all_candidates"] else 0),
                help="How to resolve conflicting evidence.",
            )

            st.markdown("#### UI Metadata")
            ui_a1, ui_a2, ui_a3, ui_a4 = st.columns(4)
            short_label = ui_a1.text_input("Short Label", value=str(ui_row.get("short_label", "") or ""), help="Compact label for dense views.")
            prefix = ui_a2.text_input("Prefix", value=str(ui_row.get("prefix", "") or ""), help="Optional display prefix.")
            suffix = ui_a3.text_input("Suffix", value=str(ui_row.get("suffix", "") or ""), help="Optional display suffix.")
            placeholder = ui_a4.text_input("Placeholder", value=str(ui_row.get("placeholder", "unk") or "unk"), help="Fallback placeholder text.")
            ui_b1, ui_b2, ui_b3 = st.columns(3)
            input_control = ui_b1.selectbox(
                "Input Control",
                ["text", "number", "toggle", "select", "multiselect", "range", "component_picker"],
                index=max(
                    0,
                    ["text", "number", "toggle", "select", "multiselect", "range", "component_picker"].index(
                        str(ui_row.get("input_control", "text"))
                        if str(ui_row.get("input_control", "text")) in ["text", "number", "toggle", "select", "multiselect", "range", "component_picker"]
                        else "text"
                    ),
                ),
                help="Preferred input control for manual review/edit tooling.",
            )
            display_decimals = ui_b2.number_input("Display Decimals", min_value=0, max_value=6, value=int(ui_row.get("display_decimals", 0) or 0), step=1, help="Numeric display precision.")
            array_handling = ui_b3.selectbox(
                "Array Handling",
                ["none", "min", "max"],
                index=max(
                    0,
                    ["none", "min", "max"].index(
                        str(rule.get("array_handling", ui_row.get("array_handling", "none")))
                        if str(rule.get("array_handling", ui_row.get("array_handling", "none"))) in ["none", "min", "max"]
                        else "none"
                    ),
                ),
                help="How array values should be represented in compact contexts.",
            )
            examples_text = st.text_input("Examples (csv)", value=csv_join(ui_row.get("examples", [])), help="Optional curated examples for this field.")
            guidance_md = st.text_area("Guidance Markdown", value=str(ui_row.get("guidance_md", "") or ""), height=90, help="Optional long-form guidance for operators.")
            deprecated_keys_text = st.text_input("Deprecated Keys (csv)", value=csv_join(rule.get("deprecated_keys", [])), help="Legacy keys that map to this contract.")

            st.markdown("#### Advanced JSON Overrides")
            selection_policy_text = st.text_area(
                "Selection Policy JSON",
                value=json.dumps(selection_policy_block or {}, ensure_ascii=False, indent=2),
                height=120,
                help="Optional deterministic candidate selection policy (for instrumented fields).",
            )
            search_hints_text = st.text_area(
                "Search Hints JSON",
                value=json.dumps(search_hints_block or {}, ensure_ascii=False, indent=2),
                height=120,
                help="Optional query/domain hint overrides for this key.",
            )
            excel_hints_text = st.text_area(
                "Excel Hints JSON",
                value=json.dumps(excel_hints_block or {}, ensure_ascii=False, indent=2),
                height=120,
                help="Optional workbook context override block stored in field_rules.",
            )

        parse_unit_conversions = parse_key_value_lines(unit_conversions_text)
        parse_token_map = parse_key_value_lines(token_map_text)
        parse_mode_aliases = parse_key_value_lines(mode_aliases_text)
        parse_unit_accepts = csv_tokens(unit_accepts_text)
        parse_delimiters = csv_tokens(delimiters_text)
        parse_range_separators = csv_tokens(range_separators_text)
        parse_accepted_formats = csv_tokens(accepted_formats_text)
        object_schema = safe_json_parse(object_schema_text, {})
        item_union = safe_json_parse(item_union_text, [])
        if not isinstance(item_union, list):
            item_union = []
        selection_policy = safe_json_parse(selection_policy_text, {})
        if not isinstance(selection_policy, dict):
            selection_policy = {}
        search_hints = safe_json_parse(search_hints_text, {})
        if not isinstance(search_hints, dict):
            search_hints = {}
        excel_hints = safe_json_parse(excel_hints_text, {})
        if not isinstance(excel_hints, dict):
            excel_hints = {}
        min_bound = None
        max_bound = None
        try:
            min_bound = None if str(range_min).strip() == "" else float(str(range_min).strip())
        except Exception:
            min_bound = None
        try:
            max_bound = None if str(range_max).strip() == "" else float(str(range_max).strip())
        except Exception:
            max_bound = None

        inferred_component_type = component_custom_type.strip() if component_type_option == "custom" else (component_type_option if component_type_option != "none" else "")
        resolved_enum_source = apply_enum_source_value(enum_source_value)
        if inferred_component_type:
            resolved_enum_source = {"type": "component_db", "ref": inferred_component_type}

        round_token = "none"
        if ftype in {"number", "integer"}:
            if int(round_decimals) <= 0:
                round_token = "int"
            elif int(round_decimals) == 1:
                round_token = "1dp"
            else:
                round_token = "2dp"

        raw_try = st.text_input("Try-It Raw Value", value="", help="Paste raw evidence text to preview normalization.")
        preview = try_normalize_preview(raw_try, parse_template, parse_unit_conversions, parse_delimiters or [",", ";", "|", "/"])
        st.code(json.dumps({"normalized_preview": preview}, ensure_ascii=False, indent=2), language="json")

        updated = deepcopy(rule)
        updated["key"] = selected_key
        updated["canonical_key"] = canonical_key.strip() or None
        updated["aliases"] = csv_tokens(aliases) or []
        updated["deprecated_keys"] = csv_tokens(deprecated_keys_text) or []
        updated["required_level"] = required_level
        updated["availability"] = availability
        updated["difficulty"] = difficulty
        updated["effort"] = int(effort)
        updated["type"] = ftype
        updated["shape"] = shape
        updated["value_form"] = value_form
        updated["unit"] = unit.strip()
        updated["parse_template"] = parse_template
        updated["enum_policy"] = enum_policy
        updated["enum_source"] = resolved_enum_source
        updated["strict_unit_required"] = bool(strict_unit_required)
        updated["round"] = round_token
        updated["array_handling"] = array_handling
        updated["min_evidence_refs"] = int(min_refs)
        updated["evidence_required"] = bool(evidence_required)
        updated["publish_gate"] = bool(publish_gate)
        updated["block_publish_when_unk"] = bool(block_publish_when_unk)
        updated["selection_policy"] = selection_policy if selection_policy else None
        updated["search_hints"] = search_hints if search_hints else {}
        updated["excel_hints"] = excel_hints if excel_hints else {}
        updated["_edited"] = True
        updated["new_value_policy"] = {
            "accept_if_evidence": bool(nvp_accept),
            "mark_needs_curation": bool(nvp_curation),
            "suggestion_target": str(new_value_target or "_suggestions/enums.json"),
        }
        updated["enum_match_strategy"] = enum_match_strategy
        updated["enum_fuzzy_threshold"] = float(enum_fuzzy_threshold)
        updated["list_rules"] = {
            "dedupe": bool(list_dedupe),
            "sort": str(list_sort),
            "min_items": int(list_min_items),
            "max_items": int(list_max_items),
        }
        updated["object_schema"] = object_schema if isinstance(object_schema, dict) else {}
        updated["item_union"] = item_union
        updated["parse_rules"] = {
            "unit": unit.strip(),
            "unit_accepts": parse_unit_accepts,
            "unit_conversions": parse_unit_conversions,
            "delimiters": parse_delimiters,
            "separators": parse_range_separators,
            "token_map": parse_token_map,
            "accepted_formats": parse_accepted_formats,
            "allow_unitless": bool(allow_unitless),
            "allow_ranges": bool(allow_ranges),
            "mode_aliases": parse_mode_aliases,
            "accept_bare_numbers_as_mode": bool(accept_bare_mode),
        }
        updated["enum"] = {
            **(updated.get("enum", {}) if isinstance(updated.get("enum"), dict) else {}),
            "policy": enum_policy,
            "source": (enum_source_ref_to_string(updated.get("enum_source")) if isinstance(updated.get("enum_source"), dict) else None),
            "match": {
                "strategy": enum_match_strategy,
                "fuzzy_threshold": float(enum_fuzzy_threshold),
            },
            "new_value_policy": updated["new_value_policy"],
        }
        updated["component"] = (
            {
                "type": inferred_component_type,
                "source": f"component_db_sources:{inferred_component_type}",
                "require_identity_evidence": bool(component_require_identity),
                "allow_new_components": bool(component_allow_new),
            }
            if inferred_component_type
            else None
        )
        updated["ui"] = {
            **(updated.get("ui", {}) if isinstance(updated.get("ui"), dict) else {}),
            "label": label.strip() or selected_key,
            "group": group.strip() or "General",
            "order": int(order),
            "tooltip_key": tooltip_key.strip() or None,
            "tooltip_md": tooltip_md,
            "tooltip_source": str(tooltip_entries.get(tooltip_key, {}).get("source", "") or "") or None,
            "display_mode": display_mode,
            "display_decimals": int(display_decimals),
            "array_handling": array_handling,
            "short_label": short_label.strip() or None,
            "prefix": prefix.strip() or None,
            "suffix": suffix.strip() or None,
            "placeholder": placeholder.strip() or "unk",
            "input_control": input_control.strip() or "text",
            "examples": csv_tokens(examples_text) or [],
            "guidance_md": guidance_md,
        }
        updated["priority"] = {
            "required_level": required_level,
            "availability": availability,
            "difficulty": difficulty,
            "effort": int(effort),
            "publish_gate": bool(publish_gate),
            "block_publish_when_unk": bool(block_publish_when_unk),
        }
        updated["contract"] = {
            **(updated.get("contract", {}) if isinstance(updated.get("contract"), dict) else {}),
            "type": ftype,
            "shape": shape,
            "unit": unit.strip(),
            "rounding": {
                "decimals": int(round_decimals),
                "mode": str(round_mode),
            },
            "range": {
                **({"min": min_bound} if min_bound is not None else {}),
                **({"max": max_bound} if max_bound is not None else {}),
            },
            "list_rules": {
                "dedupe": bool(list_dedupe),
                "sort": str(list_sort),
                "min_items": int(list_min_items),
                "max_items": int(list_max_items),
            },
            "object_schema": object_schema if isinstance(object_schema, dict) else {},
            "item_union": item_union,
            "unknown_token": "unk",
            "unknown_reason_required": True,
        }
        updated["parse"] = {
            **(updated.get("parse", {}) if isinstance(updated.get("parse"), dict) else {}),
            "template": parse_template,
            "strict_unit_required": bool(strict_unit_required),
            "unit": unit.strip(),
            "unit_accepts": parse_unit_accepts,
            "unit_conversions": parse_unit_conversions,
            "delimiters": parse_delimiters,
            "range_separators": parse_range_separators,
            "token_map": parse_token_map,
            "accepted_formats": parse_accepted_formats,
            "allow_unitless": bool(allow_unitless),
            "allow_ranges": bool(allow_ranges),
            "mode_aliases": parse_mode_aliases,
            "accept_bare_numbers_as_mode": bool(accept_bare_mode),
        }
        updated["evidence"] = {
            **(updated.get("evidence", {}) if isinstance(updated.get("evidence"), dict) else {}),
            "required": bool(evidence_required),
            "min_evidence_refs": int(min_refs),
            "tier_preference": evidence_tier_pref or ["tier1", "tier2", "tier3"],
            "conflict_policy": str(conflict_policy),
        }
        changed = json.dumps(rule, sort_keys=True, ensure_ascii=False) != json.dumps(updated, sort_keys=True, ensure_ascii=False)
        if changed and bool(st.session_state.get(autosave_key, True)):
            field_rows[selected_key] = updated
            field_rules_draft["fields"] = field_rows
            upsert_ui_catalog_row(ui_field_catalog_draft, selected_key, updated["ui"], updated)
            save_studio_drafts(category, field_rules_draft, ui_field_catalog_draft)
            st.success(f"Auto-saved `{selected_key}`.")
        elif not bool(st.session_state.get(autosave_key, True)):
            if st.button("Save Active Key", use_container_width=True, key=f"studio_save_key_{category}_{selected_key}"):
                field_rows[selected_key] = updated
                field_rules_draft["fields"] = field_rows
                upsert_ui_catalog_row(ui_field_catalog_draft, selected_key, updated["ui"], updated)
                save_studio_drafts(category, field_rules_draft, ui_field_catalog_draft)
                st.success(f"Saved `{selected_key}`.")
                st.rerun()
        errs = key_contract_errors(updated)
        if errs:
            st.error("Active key has blocking issues.")
            st.dataframe([{"error": row} for row in errs], hide_index=True, use_container_width=True)
        else:
            st.success("Active key contract is valid.")

    with tab_context:
        st.markdown(f"### Workbook Context: `{selected_key}`")
        hints = rule.get("excel_hints", {}) if isinstance(rule.get("excel_hints"), dict) else {}
        samples_raw = []
        if isinstance(hints.get("sample_values"), list):
            samples_raw.extend(hints.get("sample_values", []))
        for hint_block in hints.values():
            if isinstance(hint_block, dict) and isinstance(hint_block.get("sample_values"), list):
                samples_raw.extend(hint_block.get("sample_values", []))
        include_unknown_key = f"studio_context_include_unknown_{category}_{selected_key}"
        sample_limit_key = f"studio_context_sample_limit_{category}_{selected_key}"
        s1, s2, s3 = st.columns([1.2, 1.2, 2.6])
        include_unknown = s1.toggle("Include unknown tokens", value=bool(st.session_state.get(include_unknown_key, False)), key=include_unknown_key)
        sample_limit = int(
            s2.number_input(
                "Unique sample cap",
                min_value=1,
                max_value=50,
                value=int(st.session_state.get(sample_limit_key, 10) or 10),
                step=1,
                key=sample_limit_key,
                help="Default is first 10 unique values.",
            )
        )
        workbook_path_for_context = str(st.session_state.get(f"studio_workbook_path_{category}", workbook_map.get("workbook_path", "") or "")).strip()
        data_entry_hints = hints.get("dataEntry", {}) if isinstance(hints.get("dataEntry"), dict) else {}
        key_row_hint = int(data_entry_hints.get("row", 0) or 0)
        if key_row_hint <= 0:
            key_row_hint = find_key_row_hint(workbook_map, selected_key, fallback_row=0)
        workbook_context = workbook_context_samples(
            workbook_path=workbook_path_for_context,
            workbook_map_json=json.dumps(workbook_map or {}, ensure_ascii=False, sort_keys=True),
            selected_key=selected_key,
            key_row_hint=int(key_row_hint or 0),
            include_unknown=bool(include_unknown),
            max_unique=int(sample_limit),
            max_scan=100,
            max_blank_streak=20,
        )
        if isinstance(workbook_context, dict) and isinstance(workbook_context.get("samples"), list) and workbook_context.get("samples"):
            samples = [str(value or "").strip() for value in workbook_context.get("samples", []) if str(value or "").strip()]
            s3.caption(
                f"Source: workbook `{workbook_context.get('sheet', '?')}` "
                f"(layout `{workbook_context.get('layout', 'matrix')}`, scanned {int(workbook_context.get('scanned', 0) or 0)} columns/rows)."
            )
        else:
            fallback_samples = collect_unique_samples(
                raw_values=samples_raw,
                include_unknown=bool(include_unknown),
                max_unique=int(sample_limit),
                max_scan=100,
                max_blank_streak=20,
            )
            samples = fallback_samples.get("samples", [])
            fallback_error = str(workbook_context.get("error", "") or "")
            if fallback_error:
                s3.caption(f"Source fallback: existing hints ({fallback_error})")
            else:
                s3.caption("Source fallback: existing hint samples")
        hint_summary = infer_key_hints(selected_key, samples)

        st.markdown("#### A) Sample Values")
        if samples:
            badges = []
            if any(re.match(r"^-?\d+(\.\d+)?$", sample.lower()) for sample in samples):
                badges.append("numeric")
            if any(any(sep in sample for sep in [",", ";", "|", "/"]) for sample in samples):
                badges.append("list")
            if any(re.search(r"\d+\s*[-to]{1,3}\s*\d+", sample.lower()) for sample in samples):
                badges.append("range")
            if any(sample.lower().startswith("http://") or sample.lower().startswith("https://") for sample in samples):
                badges.append("url")
            if any(sample.lower() in {"yes", "no", "true", "false", "0", "1"} for sample in samples):
                badges.append("yes/no tokens")
            if badges:
                st.caption("Patterns: " + " | ".join([f"`{badge}`" for badge in badges]))
            st.dataframe([{"sample": sample} for sample in samples], hide_index=True, use_container_width=True, height=220)
        else:
            st.info("No sample values available.")

        st.markdown("#### B) Suggested Rule Inference")
        i1, i2, i3, i4, i5 = st.columns(5)
        i1.metric("Type", hint_summary.get("type", "string"))
        i2.metric("Shape", hint_summary.get("shape", "scalar"))
        i3.metric("Value Form", hint_summary.get("value_form", "scalar"))
        i4.metric("Unit", hint_summary.get("unit", "") or "none")
        i5.metric("Parse Template", hint_summary.get("parse_template", "text_field"))
        if st.button("Apply Suggestions", key=f"studio_apply_hint_{category}_{selected_key}", use_container_width=True):
            tmp = field_rows.get(selected_key, {})
            suggested_type = hint_summary.get("type", "string")
            suggested_shape = hint_summary.get("shape", "scalar")
            suggested_value_form = hint_summary.get("value_form", "scalar")
            suggested_unit = hint_summary.get("unit", "")
            suggested_parse = hint_summary.get("parse_template", "text_field")
            tmp["type"] = suggested_type
            tmp["shape"] = suggested_shape
            tmp["value_form"] = suggested_value_form
            tmp["unit"] = suggested_unit
            tmp["parse_template"] = suggested_parse
            contract_block = tmp.get("contract", {}) if isinstance(tmp.get("contract"), dict) else {}
            contract_block["type"] = suggested_type
            contract_block["shape"] = suggested_shape
            contract_block["unit"] = suggested_unit
            tmp["contract"] = contract_block
            parse_block_now = tmp.get("parse", {}) if isinstance(tmp.get("parse"), dict) else {}
            parse_block_now["template"] = suggested_parse
            parse_block_now["unit"] = suggested_unit
            if suggested_type in {"number", "integer"}:
                tmp["round"] = "int"
                tmp["strict_unit_required"] = bool(suggested_unit)
                parse_block_now["strict_unit_required"] = bool(suggested_unit)
            tmp["parse"] = parse_block_now
            field_rows[selected_key] = tmp
            field_rules_draft["fields"] = field_rows
            upsert_ui_catalog_row(ui_field_catalog_draft, selected_key, tmp.get("ui", {}), tmp)
            save_studio_drafts(category, field_rules_draft, ui_field_catalog_draft)
            st.success(f"Applied suggestions to `{selected_key}`.")
            st.rerun()

        st.markdown("#### C) Bind Sources")
        rule_contract_now = rule_contract(rule)
        selected_type = str(rule_contract_now.get("type", "") or "").strip().lower()
        is_numeric_key = selected_type in {"number", "integer"} or hint_summary.get("type") in {"number", "integer"}
        selected_token = token(selected_key)
        component_rows_lookup = {}
        for component_row in normalize_component_rows(workbook_map):
            ctype = str(component_row.get("component_type", "") or "").strip()
            if ctype:
                component_rows_lookup[ctype] = component_row

        component_candidates = []
        is_component_key = any(tag in selected_token for tag in ["sensor", "switch", "encoder", "material", "mcu"])
        if is_component_key:
            for ctype in component_types:
                ctoken = token(ctype)
                if ctoken and (ctoken == selected_token or ctoken in selected_token or selected_token in ctoken):
                    component_candidates.append(ctype)
            for preferred in ["sensor", "switch", "encoder", "material", "mcu"]:
                if preferred in component_types and preferred not in component_candidates and preferred in selected_token:
                    component_candidates.insert(0, preferred)
            if not component_candidates:
                component_candidates = [ctype for ctype in component_types if token(ctype) in {"sensor", "switch", "encoder", "material", "mcu"}]

        if component_candidates:
            st.caption("Component source suggestions (shown first for component-like keys)")
            for idx, ctype in enumerate(component_candidates[:6]):
                meta = component_source_meta.get(ctype, {}) if isinstance(component_source_meta.get(ctype), dict) else {}
                excel_meta = meta.get("excel", {}) if isinstance(meta.get("excel"), dict) else {}
                if not excel_meta:
                    excel_meta = {
                        "sheet": meta.get("sheet", ""),
                        "canonical_name_column": meta.get("canonical_name_column", meta.get("name_column", "")),
                    }
                if not excel_meta and ctype in component_rows_lookup:
                    fallback = component_rows_lookup.get(ctype, {})
                    excel_meta = {
                        "sheet": fallback.get("sheet", ""),
                        "canonical_name_column": fallback.get("canonical_name_column", ""),
                    }
                source_sheet = str(excel_meta.get("sheet", "") or "")
                name_col = str(excel_meta.get("canonical_name_column", "") or excel_meta.get("name_column", "") or "")
                sample_entities = meta.get("sample_entities", []) if isinstance(meta.get("sample_entities"), list) else []
                entity_count = int(meta.get("entity_count", 0) or 0)
                fallback = component_db_fallback(ctype)
                if not sample_entities:
                    sample_entities = fallback.get("sample_entities", [])
                if entity_count <= 0:
                    entity_count = int(fallback.get("entity_count", 0) or 0)
                if entity_count <= 0:
                    entity_count = len(sample_entities)
                sample_preview = ", ".join([str(v) for v in sample_entities[:10]])
                st.caption(f"{ctype} (sheet:{source_sheet or '?'} name_col:{name_col or '?'}) count:{entity_count} samples:{sample_preview or 'n/a'}")
                if st.button(
                    f"Bind Component {ctype}",
                    key=f"studio_context_bind_component_{category}_{selected_key}_{idx}",
                    use_container_width=True,
                ):
                    tmp = field_rows.get(selected_key, {})
                    enum_obj = tmp.get("enum", {}) if isinstance(tmp.get("enum"), dict) else {}
                    enum_policy = str(tmp.get("enum_policy", "") or enum_obj.get("policy", "") or "").strip().lower()
                    if not enum_policy:
                        enum_policy = "open_prefer_known"
                    tmp["enum_policy"] = enum_policy
                    tmp["enum_source"] = {"type": "component_db", "ref": ctype}
                    enum_obj["policy"] = enum_policy
                    enum_obj["source"] = f"component_db_sources:{ctype}"
                    tmp["enum"] = enum_obj
                    tmp["parse_template"] = "component_reference"
                    tmp["component"] = {
                        "type": ctype,
                        "source": f"component_db_sources:{ctype}",
                        "require_identity_evidence": True,
                        "allow_new_components": True,
                    }
                    field_rows[selected_key] = tmp
                    field_rules_draft["fields"] = field_rows
                    upsert_ui_catalog_row(ui_field_catalog_draft, selected_key, tmp.get("ui", {}), tmp)
                    save_studio_drafts(category, field_rules_draft, ui_field_catalog_draft)
                    st.success(f"Bound component source `{ctype}`.")
                    st.rerun()

        if is_numeric_key:
            st.info("Numeric key detected: enum bucket suggestions are hidden by default.")
        elif component_candidates:
            st.caption("Enum suggestions are deprioritized for component-like keys.")
        else:
            bucket_candidates = []
            for bucket in enum_fields:
                bucket_token = token(bucket)
                if bucket_token == selected_token or selected_token in bucket_token or bucket_token in selected_token:
                    bucket_candidates.append(bucket)
            if not bucket_candidates and selected_token:
                bucket_candidates = [bucket for bucket in enum_fields if token(bucket).startswith(selected_token[:4])][:6]
            if bucket_candidates:
                st.caption("Enum bucket suggestions")
                for idx, bucket in enumerate(bucket_candidates[:8]):
                    meta = enum_bucket_meta.get(bucket, {}) if isinstance(enum_bucket_meta.get(bucket), dict) else {}
                    excel_meta = meta.get("excel", {}) if isinstance(meta.get("excel"), dict) else {}
                    values = meta.get("values", []) if isinstance(meta.get("values"), list) else []
                    source_sheet = str(excel_meta.get("sheet", meta.get("sheet", "")) or "")
                    source_col = str(excel_meta.get("column", "") or "")
                    if not source_col and meta.get("column_index"):
                        source_col = index_to_col(int(meta.get("column_index") or 0))
                    header = str(excel_meta.get("header", meta.get("header", bucket)) or bucket)
                    sample_preview = ", ".join([str(v) for v in values[:10]])
                    st.caption(f"{bucket} (sheet:{source_sheet or '?'} column:{source_col or '?'} header:{header}) count:{len(values)} samples:{sample_preview or 'n/a'}")
                    if st.button(
                        f"Bind Enum {bucket}",
                        key=f"studio_context_bind_enum_{category}_{selected_key}_{idx}",
                        use_container_width=True,
                    ):
                        tmp = field_rows.get(selected_key, {})
                        enum_obj = tmp.get("enum", {}) if isinstance(tmp.get("enum"), dict) else {}
                        enum_policy = str(tmp.get("enum_policy", "") or enum_obj.get("policy", "") or "").strip().lower()
                        if not enum_policy:
                            enum_policy = "open_prefer_known"
                        tmp["enum_policy"] = enum_policy
                        tmp["enum_source"] = {"type": "known_values", "ref": bucket}
                        enum_obj["policy"] = enum_policy
                        enum_obj["source"] = f"enum_buckets:{bucket}"
                        tmp["enum"] = enum_obj
                        field_rows[selected_key] = tmp
                        field_rules_draft["fields"] = field_rows
                        upsert_ui_catalog_row(ui_field_catalog_draft, selected_key, tmp.get("ui", {}), tmp)
                        save_studio_drafts(category, field_rules_draft, ui_field_catalog_draft)
                        st.success(f"Bound enum bucket `{bucket}`.")
                        st.rerun()
            else:
                st.info("No enum bucket suggestions for this key.")

    with tab_reports:
        if compile_errors:
            st.error("Compile has blocking errors.")
            st.dataframe([{"error": row} for row in compile_errors], use_container_width=True, height=260)
        if compile_warnings:
            st.warning("Compile warnings")
            st.dataframe([{"warning": row} for row in compile_warnings], use_container_width=True, height=220)
        with st.expander("Artifact Paths", expanded=False):
            st.code(
                "\n".join(
                    [
                        str(paths["workbook_map"]),
                        str(paths["field_rules_draft"]),
                        str(paths["ui_field_catalog_draft"]),
                        str(paths["field_rules_full"]),
                        str(paths["field_rules"]),
                        str(paths["field_rules_runtime"]),
                        str(paths["ui_field_catalog"]),
                        str(paths["known_values"]),
                        str(paths["compile_report"]),
                    ]
                ),
                language="text",
            )
        with st.expander("Compile Report JSON", expanded=False):
            st.json(compile_report)

def run_root_paths(category: str, pid: str, run_id: str):
    return {
        "run": RUNS_ROOT / category / pid / run_id,
    }


def read_run_bundle(category: str, pid: str, run_id: str):
    if not category or not pid or not run_id:
        return {}
    roots = run_root_paths(category, pid, run_id)
    summary_compact = read_json(roots["run"] / "summary.json", {})
    spec = read_json(roots["run"] / "spec.json", {})
    provenance = read_json(roots["run"] / "provenance.json", {})
    traffic_light = read_json(roots["run"] / "traffic_light.json", {})
    return {
        "exists": bool(summary_compact or spec),
        "run_id": run_id,
        "summary": summary_compact,
        "summary_compact": summary_compact,
        "summary_detailed": {},
        "spec": spec,
        "provenance": provenance,
        "traffic_light": traffic_light,
        "paths": roots,
    }


def normalize_summary(summary):
    return summary if isinstance(summary, dict) else {}


def schema_field_order(schema: dict):
    field_order = list((schema or {}).get("field_order", []) or [])
    editorial = set((schema or {}).get("editorial_fields", []) or [])
    return [field for field in field_order if field not in editorial]


def live_field_decisions(events, allowed_fields):
    out = {}
    allowed = set(allowed_fields or [])
    for event in events:
        if str(event.get("event", "")) != "field_decision":
            continue
        field = str(event.get("field", "")).strip()
        if not field:
            continue
        if allowed and field not in allowed:
            continue
        out[field] = {
            "value": event.get("value", "unk"),
            "decision": event.get("decision", ""),
            "unknown_reason": event.get("unknown_reason", ""),
            "confidence": float(event.get("confidence", 0) or 0),
            "traffic_color": str(event.get("traffic_color", "")).lower(),
            "traffic_reason": event.get("traffic_reason", ""),
            "evidence_count": int(event.get("evidence_count", 0) or 0),
        }
    return out


def choose_display_value(field, spec_row, live_row):
    if known(spec_row):
        return spec_row
    if live_row and known(live_row.get("value")):
        return live_row.get("value")
    return spec_row


def latest_event(events, event_name: str):
    for event in reversed(events or []):
        if str(event.get("event", "")) == event_name:
            return event
    return {}


def build_live_helper_llm_snapshot(events):
    helper_loaded = latest_event(events, "helper_files_context_loaded")
    helper_fill = latest_event(events, "helper_supportive_fill_applied")
    budget_block = latest_event(events, "llm_extract_skipped_budget")
    llm_started_events = [event for event in (events or []) if str(event.get("event", "")) == "llm_call_started"]
    llm_usage_events = [event for event in (events or []) if str(event.get("event", "")) == "llm_call_usage"]

    provider = "n/a"
    if llm_usage_events:
        provider = str(llm_usage_events[-1].get("provider", "")).strip() or provider
    if provider == "n/a" and llm_started_events:
        provider = str(llm_started_events[-1].get("provider", "")).strip() or provider

    live_helper = {
        "active_filtering_match": bool(helper_loaded.get("active_match")),
        "supportive_match_count": int(helper_loaded.get("supportive_matches", 0) or 0),
        "supportive_file_count": int(helper_loaded.get("supportive_files", 0) or 0),
        "seed_urls_from_active_count": int(helper_loaded.get("helper_seed_urls", 0) or 0),
        "supportive_fields_filled_count": int(helper_fill.get("fields_filled", 0) or 0),
        "supportive_fields_filled_by_method": helper_fill.get("fields_filled_by_method", {}) or {},
    }
    live_llm = {
        "provider": provider,
        "call_count_run": int(len(llm_started_events)),
        "cost_usd_run": float(sum(float(event.get("cost_usd", 0) or 0) for event in llm_usage_events)),
        "budget": {
            "blocked_reason": str(budget_block.get("reason", "")).strip() or None
        }
    }
    return live_helper, live_llm


def merge_display_helper_info(summary_helper, live_helper):
    base = dict(summary_helper or {})
    live = dict(live_helper or {})
    if "active_filtering_match" not in base:
        base["active_filtering_match"] = live.get("active_filtering_match", False)
    base["supportive_match_count"] = max(
        int(base.get("supportive_match_count", 0) or 0),
        int(live.get("supportive_match_count", 0) or 0),
    )
    base["supportive_fields_filled_count"] = max(
        int(base.get("supportive_fields_filled_count", 0) or 0),
        int(live.get("supportive_fields_filled_count", 0) or 0),
    )
    if not base.get("supportive_fields_filled_by_method") and live.get("supportive_fields_filled_by_method"):
        base["supportive_fields_filled_by_method"] = live.get("supportive_fields_filled_by_method")
    return base


def merge_display_llm_info(summary_llm, live_llm):
    base = dict(summary_llm or {})
    live = dict(live_llm or {})
    if not str(base.get("provider", "")).strip() or str(base.get("provider", "")).strip().lower() in {"n/a", "unknown"}:
        base["provider"] = live.get("provider", "n/a")
    base["call_count_run"] = max(
        int(base.get("call_count_run", 0) or 0),
        int(live.get("call_count_run", 0) or 0),
    )
    base["cost_usd_run"] = max(
        float(base.get("cost_usd_run", 0) or 0),
        float(live.get("cost_usd_run", 0) or 0),
    )
    budget = dict(base.get("budget", {}) or {})
    if not budget.get("blocked_reason"):
        live_reason = ((live.get("budget", {}) or {}).get("blocked_reason"))
        if live_reason:
            budget["blocked_reason"] = live_reason
    if budget:
        base["budget"] = budget
    return base


@st.cache_data(ttl=30, show_spinner=False)
def list_categories():
    names = set()
    if HELPER_ROOT.exists():
        for item in HELPER_ROOT.iterdir():
            if item.is_dir():
                names.add(item.name)
    return sorted(names)


@st.cache_data(ttl=30, show_spinner=False)
def load_schema(category: str):
    generated_root = HELPER_ROOT / category / "_generated"
    generated_schema = read_json(generated_root / "schema.json", {})
    if isinstance(generated_schema, dict) and generated_schema:
        return generated_schema

    field_rules = read_json(generated_root / "field_rules.json", {})
    ui_catalog = read_json(generated_root / "ui_field_catalog.json", {})
    fields = field_rules.get("fields") if isinstance(field_rules, dict) else {}
    if not isinstance(fields, dict) or not fields:
        return {}

    ui_order = {}
    rows = ui_catalog.get("fields") if isinstance(ui_catalog, dict) else []
    if isinstance(rows, list):
        for row in rows:
            if not isinstance(row, dict):
                continue
            key = str(row.get("key", "")).strip()
            if not key:
                continue
            try:
                ui_order[key] = int(row.get("order", 10_000_000))
            except Exception:
                ui_order[key] = 10_000_000

    def _norm(value):
        return "".join(ch.lower() if ch.isalnum() or ch == "_" else "_" for ch in str(value or "")).strip("_")

    entries = []
    for raw_key, raw_rule in fields.items():
        key = _norm(raw_key)
        if not key or not isinstance(raw_rule, dict):
            continue
        entries.append((key, raw_rule))
    entries.sort(key=lambda item: (ui_order.get(item[0], 10_000_000), item[0]))

    critical = []
    expected_easy = []
    expected_sometimes = []
    deep = []
    order = []
    for key, rule in entries:
        order.append(key)
        priority = rule_priority(rule)
        required_level = str(priority.get("required_level", "")).strip().lower()
        availability = str(priority.get("availability", "")).strip().lower()
        difficulty = str(priority.get("difficulty", "")).strip().lower()
        if required_level == "critical":
            critical.append(key)
        if required_level in {"required", "critical", "expected"}:
            if difficulty == "easy" or availability == "expected":
                expected_easy.append(key)
            else:
                expected_sometimes.append(key)
        else:
            deep.append(key)

    return {
        "category": category,
        "field_order": order,
        "critical_fields": sorted(set(critical)),
        "expected_easy_fields": sorted(set(expected_easy)),
        "expected_sometimes_fields": sorted(set(expected_sometimes)),
        "deep_fields": sorted(set(deep)),
        "editorial_fields": []
    }


@st.cache_data(ttl=30, show_spinner=False)
def load_active_filtering(category: str):
    path = HELPER_ROOT / category / "activeFiltering.json"
    payload = read_json(path, [])
    out = []
    if not isinstance(payload, list):
        return out
    for idx, row in enumerate(payload):
        if not isinstance(row, dict):
            continue
        brand = norm(row.get("brand"))
        model = norm(row.get("model"))
        if not brand or not model:
            continue
        out.append(
            {
                "brand": brand,
                "model": model,
                "variant": clean_variant(row.get("variant")),
                "id": row.get("id", idx),
                "raw": row,
            }
        )
    return out


@st.cache_data(ttl=60, show_spinner=False)
def load_supportive(category: str):
    return {"files": {}, "records": []}


@st.cache_data(ttl=10, show_spinner=False)
def list_final_rows(category: str):
    rows = []
    root = FINAL_ROOT / category
    if not root.exists():
        return rows
    for brand_dir in root.iterdir():
        if not brand_dir.is_dir():
            continue
        for model_dir in brand_dir.iterdir():
            if not model_dir.is_dir():
                continue
            candidates = [model_dir] + [d for d in model_dir.iterdir() if d.is_dir()]
            for path in candidates:
                summary_path = path / "summary.json"
                if not summary_path.exists():
                    continue
                summary = read_json(summary_path, {})
                meta = read_json(path / "meta.json", {})
                identity = (meta or {}).get("canonical_identity", {})
                rows.append(
                    {
                        "brand": norm(identity.get("brand")) or brand_dir.name,
                        "model": norm(identity.get("model")) or model_dir.name,
                        "variant": clean_variant(identity.get("variant")) or (path.name if path != model_dir else ""),
                        "path": path,
                        "summary": summary,
                    }
                )
    return rows


@st.cache_data(ttl=5, show_spinner=False)
def load_queue(category: str):
    return read_json(effective_queue_root() / category / "state.json", {"products": {}})


@st.cache_data(ttl=5, show_spinner=False)
def load_billing_month():
    month = datetime.now(timezone.utc).strftime("%Y-%m")
    return read_json(effective_billing_root() / "monthly" / f"{month}.json", {})


@st.cache_data(ttl=15, show_spinner=False)
def load_field_availability(category: str):
    return read_json(effective_learning_root() / category / "field_availability.json", {})


def product_id(category: str, brand: str, model: str, variant: str):
    parts = [category, brand, model]
    if clean_variant(variant):
        parts.append(clean_variant(variant))
    return slug("-".join(parts))


def filter_events_for_product(events, category: str, pid: str):
    filtered = []
    for event in events:
        if str(event.get("category", "")) != category:
            continue
        event_pid = str(event.get("productId", ""))
        if event_pid != pid:
            continue
        filtered.append(event)
    return filtered


def latest_run_id(events, pid: str = ""):
    normalized_pid = str(pid or "")
    if normalized_pid:
        for event in reversed(events):
            if str(event.get("productId", "")) != normalized_pid:
                continue
            run_id = str(event.get("runId", "")).strip()
            if run_id:
                return run_id
    for event in reversed(events):
        run_id = str(event.get("runId", "")).strip()
        if run_id:
            return run_id
    return ""


def filter_events_for_run(events, run_id: str):
    if not run_id:
        return events
    keep = []
    for event in events:
        value = str(event.get("runId", "")).strip()
        if value in {"", run_id}:
            keep.append(event)
    return keep


def pipeline_stage_status(events):
    counts = Counter(str(event.get("event", "")) for event in events)
    done = []
    weights = [10, 12, 12, 16, 16, 16, 9, 9]
    score = 0.0
    for idx, (stage_key, label, stage_events) in enumerate(PIPELINE_STAGE_DEFS):
        reached = any(counts.get(event_name, 0) > 0 for event_name in stage_events)
        done.append(
            {
                "stage_key": stage_key,
                "label": label,
                "reached": reached,
                "events": sorted(stage_events),
            }
        )
        if reached:
            score += weights[idx]

    score = min(100.0, max(0.0, score))
    return {
        "stages": done,
        "counts": counts,
        "progress_pct": score
    }


def event_rows_with_help(counts: Counter):
    rows = []
    for event_name, count in counts.most_common():
        rows.append(
            {
                "event": event_name,
                "count": int(count),
                "meaning": EVENT_MEANINGS.get(event_name, "No tooltip defined yet."),
            }
        )
    return rows


def build_catalog(category: str):
    active = load_active_filtering(category)
    finals = list_final_rows(category)
    queue = load_queue(category)

    rows = {}

    def put(brand, model, variant):
        key = (token(brand), token(model), token(variant))
        if key not in rows:
            rows[key] = {
                "brand": brand,
                "model": model,
                "variant": variant,
                "in_active": False,
                "has_final": False,
                "validated": False,
                "confidence": 0.0,
                "completeness_required": 0.0,
                "final_path": "",
                "queue_status": "",
            }
        return rows[key]

    for row in active:
        entry = put(row["brand"], row["model"], row["variant"])
        entry["in_active"] = True

    for row in finals:
        entry = put(row["brand"], row["model"], row["variant"])
        entry["has_final"] = True
        entry["validated"] = bool(row["summary"].get("validated", False))
        entry["confidence"] = float(row["summary"].get("confidence", 0) or 0)
        entry["completeness_required"] = float(row["summary"].get("completeness_required", 0) or 0)
        entry["final_path"] = str(row["path"])

    qmap = {
        norm(v.get("productId")): norm(v.get("status"))
        for v in (queue.get("products", {}) or {}).values()
        if isinstance(v, dict)
    }
    for entry in rows.values():
        pid = product_id(category, entry["brand"], entry["model"], entry["variant"])
        entry["queue_status"] = qmap.get(pid, "")
        entry["product_id"] = pid

    out = sorted(rows.values(), key=lambda r: (token(r["brand"]), token(r["model"]), token(r["variant"])))
    return {"rows": out, "active": active, "supportive": {"files": {}, "records": []}, "queue": queue}


def find_bundle(category: str, brand: str, model: str, variant: str):
    base = FINAL_ROOT / slug(category) / slug(brand) / slug(model)
    candidates = []
    variant_clean = clean_variant(variant)
    if variant_clean and (base / slug(variant_clean) / "summary.json").exists():
        candidates.append(base / slug(variant_clean))
    if (base / "summary.json").exists():
        candidates.append(base)
    if base.exists():
        for child in base.iterdir():
            if child.is_dir() and (child / "summary.json").exists():
                candidates.append(child)
    if not candidates:
        return {"exists": False}
    best = sorted(
        candidates,
        key=lambda p: (
            float((read_json(p / "summary.json", {}) or {}).get("completeness_required", 0)),
            float((read_json(p / "summary.json", {}) or {}).get("confidence", 0)),
        ),
        reverse=True,
    )[0]
    meta = read_json(best / "meta.json", {})
    run_id = str((meta or {}).get("runId", "")).strip()
    pid = str((meta or {}).get("productId", "")).strip()
    debug_summary = {}
    if run_id and pid:
        run_bundle = read_run_bundle(category, pid, run_id)
        debug_summary = normalize_summary(run_bundle.get("summary"))

    return {
        "exists": True,
        "path": best,
        "spec": read_json(best / "spec.json", {}),
        "summary": read_json(best / "summary.json", {}),
        "provenance": read_json(best / "provenance.json", {}),
        "traffic_light": read_json(best / "traffic_light.json", {}),
        "meta": meta,
        "run_id": run_id,
        "product_id": pid,
        "debug_summary": debug_summary,
    }


def ensure_proc_state():
    if "proc" not in st.session_state:
        st.session_state.proc = None
    if "proc_handle" not in st.session_state:
        st.session_state.proc_handle = None
    if "proc_log" not in st.session_state:
        st.session_state.proc_log = []
    if "proc_log_path" not in st.session_state:
        st.session_state.proc_log_path = str(GUI_PROCESS_LOG_PATH)
    if "last_cmd" not in st.session_state:
        st.session_state.last_cmd = ""
    if "auto_refresh" not in st.session_state:
        st.session_state.auto_refresh = True


def poll_proc():
    proc = st.session_state.proc
    log_path = Path(st.session_state.proc_log_path)
    st.session_state.proc_log = read_tail_lines(log_path, 400)
    if not proc:
        return
    if proc.poll() is not None:
        st.session_state.proc = None
        handle = st.session_state.proc_handle
        if handle:
            try:
                handle.flush()
                handle.close()
            except Exception:
                pass
            st.session_state.proc_handle = None


def start_cli(args):
    proc = st.session_state.proc
    if proc and proc.poll() is None:
        st.warning("A process is already running. Stop it first.")
        return
    st.session_state.last_cmd = "node src/cli/spec.js " + " ".join(args)
    runtime_dir = OUTPUT_ROOT / "_runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    log_path = Path(st.session_state.proc_log_path)
    with log_path.open("w", encoding="utf-8") as log_seed:
        log_seed.write(f"$ {st.session_state.last_cmd}\n")
        log_seed.write(f"Started {datetime.now(timezone.utc).isoformat()}\n")
    try:
        handle = log_path.open("a", encoding="utf-8")
        st.session_state.proc_handle = handle
        st.session_state.proc = subprocess.Popen(
            ["node", "src/cli/spec.js", *args],
            cwd=str(REPO_ROOT),
            stdout=handle,
            stderr=subprocess.STDOUT,
            text=True,
        )
    except Exception as exc:
        st.session_state.proc = None
        if st.session_state.proc_handle:
            try:
                st.session_state.proc_handle.close()
            except Exception:
                pass
            st.session_state.proc_handle = None
        st.session_state.proc_log = [
            f"$ {st.session_state.last_cmd}",
            f"Failed to start process: {exc}",
            "Verify Node.js is installed and accessible in PATH.",
        ]


def stop_cli():
    proc = st.session_state.proc
    if proc:
        try:
            proc.terminate()
        except Exception:
            pass
        st.session_state.proc = None
    handle = st.session_state.proc_handle
    if handle:
        try:
            handle.flush()
            handle.close()
        except Exception:
            pass
        st.session_state.proc_handle = None
    st.session_state.proc_log.append("Process terminated.")


st.set_page_config(page_title="Spec Factory Control Center", layout="wide")
ensure_proc_state()
poll_proc()

categories = list_categories()
if not categories:
    st.error("No category folders found.")
    st.stop()

with st.sidebar:
    st.header("Selection")
    category = st.selectbox(
        "Category",
        categories,
        index=categories.index("mouse") if "mouse" in categories else 0,
        help="Select category from `helper_files/<category>`."
    )
    catalog = build_catalog(category)
    rows = catalog["rows"]
    if not rows:
        st.warning("No products found in helper/final data.")
        st.stop()

    active_rows = [r for r in rows if r["in_active"]] or rows
    brands = sorted({r["brand"] for r in active_rows})
    brand = st.selectbox(
        "Brand",
        brands,
        help="Brand list is merged from active targets, queue state, and final outputs."
    )
    models = sorted({r["model"] for r in active_rows if r["brand"] == brand})
    model = st.selectbox("Model", models, help="Model target from helper catalog/final results.")
    variants = sorted(
        {clean_variant(r["variant"]) for r in active_rows if r["brand"] == brand and r["model"] == model},
        key=lambda v: (v != "", token(v))
    )
    variant_label = st.selectbox(
        "Variant",
        ["(none)" if v == "" else v for v in variants],
        help="Use variant when the same model has multiple editions; improves identity validation."
    )
    variant = "" if variant_label == "(none)" else variant_label

    st.header("Run")
    mode = st.selectbox(
        "Accuracy Mode",
        ["aggressive", "balanced"],
        help="`aggressive` increases search breadth and crawl depth for higher field coverage at higher cost/time."
    )
    profile = st.selectbox(
        "Run Profile",
        ["fast", "standard", "thorough"],
        index=2,
        help="`fast` quick pass, `standard` balanced, `thorough` highest effort."
    )
    rounds = st.number_input("Max Rounds", min_value=1, max_value=20, value=2, step=1)
    local = st.checkbox("Local Mode", value=True, help="Use local-first storage paths.")
    dry_run = st.checkbox("Dry Run", value=False, help="Plan/simulate without full network extraction.")
    st.session_state.auto_refresh = st.checkbox("Auto Refresh While Running", value=st.session_state.auto_refresh)

    args = ["run-ad-hoc", category, brand, model]
    if variant:
        args.append(variant)
    args += ["--until-complete", "--mode", mode, "--max-rounds", str(int(rounds)), "--profile", profile]
    if local:
        args.append("--local")
    if dry_run:
        args.append("--dry-run")
    st.code("node src/cli/spec.js " + " ".join(args), language="bash")

    if st.button(
        "Run Selected Product",
        use_container_width=True,
        help="Runs selected product now, multi-round, until complete/exhausted/limits reached."
    ):
        start_cli(args)
    if st.button(
        "Daemon Once",
        use_container_width=True,
        help="One full scheduler cycle: ingest imports, pick queue items, run processing once."
    ):
        dargs = ["daemon", "--category", category, "--mode", mode, "--once"] + (["--local"] if local else [])
        start_cli(dargs)
    if st.button(
        "Daemon Continuous",
        use_container_width=True,
        help="Always-on loop for category queue. Use Stop Process to halt."
    ):
        dargs = ["daemon", "--category", category, "--mode", mode] + (["--local"] if local else [])
        start_cli(dargs)
    if st.button(
        "Watch Imports Once",
        use_container_width=True,
        help="Scans `imports/<category>/incoming` and converts CSV rows into product jobs once."
    ):
        iargs = ["watch-imports", "--category", category, "--once"] + (["--local"] if local else [])
        start_cli(iargs)
    if st.button("Stop Process", use_container_width=True):
        stop_cli()
    if st.button("Refresh", use_container_width=True):
        st.rerun()

    with st.expander("Run Button Help", expanded=False):
        st.markdown(
            "- `Run Selected Product`: Immediate product run with current mode/profile/round limits.\n"
            "- `Daemon Once`: Ingest + queue + run once, then exit.\n"
            "- `Daemon Continuous`: Keep running queue forever until stopped.\n"
            "- `Watch Imports Once`: Only intake CSVs to queue, no product extraction.\n"
            "- `Auto Refresh While Running`: refreshes the dashboard every ~1.5s so progress updates live."
        )

st.title("Spec Factory Control Center")
st.caption(f"Repo: `{REPO_ROOT}` | Output: `{OUTPUT_ROOT}`")

bundle = find_bundle(category, brand, model, variant)
schema = load_schema(category)
fields = schema_field_order(schema)
critical = set(schema.get("critical_fields", []))
pid = product_id(category, brand, model, variant)
events = read_jsonl(EVENTS_PATH, limit=4000)
events_for_product = filter_events_for_product(events, category, pid)
latest_selected_run = latest_run_id(events_for_product, pid=pid)
events_for_selected_run = filter_events_for_run(events_for_product, latest_selected_run)
run_bundle = read_run_bundle(category, pid, latest_selected_run)
active_summary = normalize_summary(run_bundle.get("summary"))
if not active_summary:
    active_summary = normalize_summary(bundle.get("debug_summary", {})) or normalize_summary(bundle.get("summary", {}))
active_spec = run_bundle.get("spec", {}) if run_bundle else {}
spec = active_spec if active_spec else (bundle.get("spec", {}) if bundle.get("exists") else {})
active_traffic = run_bundle.get("traffic_light", {}) if run_bundle else {}
traffic_map = (active_traffic or bundle.get("traffic_light", {}) or {}).get("by_field", {})
reasoning = normalize_summary(active_summary.get("field_reasoning", {}))
live_decisions = live_field_decisions(events_for_selected_run, fields)

tab1, tab2, tab3, tab4, tab5 = st.tabs(["Overview", "Selected Product", "Live Runtime", "Billing & Learning", "Field Rules Studio"])

with tab1:
    monthly = load_billing_month() or {}
    totals = monthly.get("totals", {})
    targets = sum(1 for r in rows if r["in_active"])
    finals = sum(1 for r in rows if r["has_final"])
    validated = sum(1 for r in rows if r["has_final"] and r["validated"])
    coverage = finals / targets if targets else 0
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Target Products", targets)
    c2.metric("Final Outputs", finals)
    c3.metric("Validated", validated)
    c4.metric("Coverage", f"{coverage * 100:.1f}%")
    c5.metric("Monthly Cost USD", f"{float(totals.get('cost_usd', 0)):.6f}")
    st.progress(coverage, text=f"Final coverage across helper targets: {coverage * 100:.1f}%")
    table = []
    for r in rows:
        state = "validated" if r["has_final"] and r["validated"] else ("partial" if r["has_final"] else (r["queue_status"] or "target_only"))
        table.append(
            {
                "brand": r["brand"],
                "model": r["model"],
                "variant": r["variant"],
                "state": state,
                "active_target": r["in_active"],
                "confidence": r["confidence"],
                "completeness_required": r["completeness_required"],
            }
        )
    st.dataframe(table, use_container_width=True, height=420)

with tab2:
    st.subheader(f"{category} / {brand} / {model}" + (f" / {variant}" if variant else ""))
    st.caption(f"Product ID: `{pid}`")
    has_final_bundle = bool(bundle.get("exists"))
    if has_final_bundle:
        st.caption(f"Final Path: `{bundle.get('path')}`")
    else:
        st.info("No final output bundle found yet. Live run status below will still update.")

    run_focus = latest_selected_run or bundle.get("run_id", "")
    running_now = bool(st.session_state.proc and st.session_state.proc.poll() is None)
    if run_focus:
        st.caption(f"Run focus: `{run_focus}`")
    if running_now:
        st.warning("A CLI process is currently running. Live fields and events update automatically.")

    summary_for_metrics = normalize_summary(active_summary or bundle.get("summary", {}))
    s1, s2, s3, s4 = st.columns(4)
    s1.metric("Validated", "yes" if summary_for_metrics.get("validated") else "no")
    s2.metric("Confidence", f"{float(summary_for_metrics.get('confidence', 0) or 0):.4f}")
    s3.metric("Completeness Required", f"{float(summary_for_metrics.get('completeness_required', 0) or 0):.4f}")
    s4.metric("Coverage Overall", f"{float(summary_for_metrics.get('coverage_overall', 0) or 0):.4f}")

    live_stage = pipeline_stage_status(events_for_selected_run)
    st.progress(
        live_stage["progress_pct"] / 100.0,
        text=f"Live run progress: {live_stage['progress_pct']:.0f}%"
    )
    live_counts = live_stage["counts"]
    p1, p2, p3, p4 = st.columns(4)
    p1.metric("Sources Fetch Started", int(live_counts.get("source_fetch_started", 0)))
    p2.metric("Sources Processed", int(live_counts.get("source_processed", 0)))
    p3.metric("LLM Calls Started", int(live_counts.get("llm_call_started", 0)))
    p4.metric("Field Decisions Logged", int(live_counts.get("field_decision", 0)))

    helper_info = summary_for_metrics.get("helper_files", {}) if isinstance(summary_for_metrics, dict) else {}
    llm_info = summary_for_metrics.get("llm", {}) if isinstance(summary_for_metrics, dict) else {}
    live_helper_info, live_llm_info = build_live_helper_llm_snapshot(events_for_selected_run)
    helper_info = merge_display_helper_info(helper_info, live_helper_info)
    llm_info = merge_display_llm_info(llm_info, live_llm_info)
    st.subheader("Helper + LLM Status")
    h1, h2, h3, h4 = st.columns(4)
    h1.metric(
        "Helper Active Match",
        "yes" if bool(helper_info.get("active_filtering_match")) else "no",
        help="True when activeFiltering brand/model matched the selected product."
    )
    h2.metric(
        "Helper Support Matches",
        int(helper_info.get("supportive_match_count", 0) or 0),
        help="Supportive helper rows matching identity."
    )
    h3.metric(
        "Helper Fields Filled",
        int(helper_info.get("supportive_fields_filled_count", 0) or 0),
        help="Fields filled directly from supportive helper evidence."
    )
    h4.metric(
        "LLM Provider",
        str(llm_info.get("provider", "n/a")),
        help="Configured provider for plan/extract/validate calls."
    )
    l1, l2, l3, l4 = st.columns(4)
    l1.metric("LLM Calls (Run)", int(llm_info.get("call_count_run", 0) or 0))
    l2.metric("LLM Cost (Run USD)", f"{float(llm_info.get('cost_usd_run', 0) or 0):.6f}")
    l3.metric(
        "LLM Failures (Run Events)",
        int(
            sum(
                1
                for event in events_for_selected_run
                if str(event.get("event", "")) in {"llm_call_failed", "llm_extract_failed"}
            )
        ),
    )
    l4.metric(
        "LLM Filled Fields",
        int(llm_info.get("fields_filled_by_llm_count", 0) or 0),
        help="Fields whose accepted provenance includes llm_extract/llm_validate."
    )
    blocked_reason = (
        ((llm_info.get("budget", {}) or {}).get("blocked_reason"))
        if isinstance(llm_info, dict)
        else ""
    )
    st.caption(f"LLM budget block reason: `{blocked_reason or 'none'}`")
    with st.expander("Helper details", expanded=False):
        st.json(helper_info)
    with st.expander("LLM details", expanded=False):
        st.json(llm_info)
    if not summary_for_metrics:
        st.caption("Detailed run summary not found yet; run a product to populate full helper/llm details.")
    helper_filled = helper_info.get("supportive_fields_filled", []) if isinstance(helper_info, dict) else []
    if helper_filled:
        st.caption("Helper-filled fields:")
        st.write(", ".join(helper_filled))

    field_rows = []
    order = fields or sorted(spec.keys())
    for key in order:
        live_row = live_decisions.get(key, {})
        value = choose_display_value(key, spec.get(key, "unk"), live_row)
        trow = traffic_map.get(key, {}) if isinstance(traffic_map, dict) else {}
        rrow = reasoning.get(key, {}) if isinstance(reasoning, dict) else {}
        unknown_reason = rrow.get("unknown_reason", "") if isinstance(rrow, dict) else ""
        if not unknown_reason and value == "unk":
            unknown_reason = live_row.get("unknown_reason", "")
        field_rows.append(
            {
                "field": key,
                "critical": key in critical,
                "status": "Collected" if known(value) else "Unknown",
                "value": value,
                "traffic": (
                    str(trow.get("color", "")).upper()
                    or str(live_row.get("traffic_color", "")).upper()
                    or "N/A"
                ),
                "unknown_reason": unknown_reason,
            }
        )
    known_count = sum(1 for row in field_rows if row["status"] == "Collected")
    crit_total = sum(1 for row in field_rows if row["critical"])
    crit_known = sum(1 for row in field_rows if row["critical"] and row["status"] == "Collected")
    if field_rows:
        p1, p2 = st.columns(2)
        p1.progress(known_count / max(len(field_rows), 1), text=f"Collected fields: {known_count}/{len(field_rows)}")
        p2.progress(crit_known / max(crit_total, 1) if crit_total else 0, text=f"Critical collected: {crit_known}/{crit_total}")
        st.dataframe(field_rows, use_container_width=True, height=520)
        st.caption(f"Unknown reason breakdown: {dict(Counter(row['unknown_reason'] for row in field_rows if row['status'] == 'Unknown'))}")

        availability_artifact = load_field_availability(category) or {}
        availability_fields = availability_artifact.get("fields", {}) if isinstance(availability_artifact, dict) else {}
        availability_rows = []
        expected_missing = 0
        sometimes_missing = 0
        rare_missing = 0
        for row in field_rows:
            f = row["field"]
            availability = availability_fields.get(f, {}) if isinstance(availability_fields, dict) else {}
            klass = str(availability.get("classification", "sometimes"))
            if row["status"] == "Unknown":
                if klass == "expected":
                    expected_missing += 1
                elif klass == "rare":
                    rare_missing += 1
                else:
                    sometimes_missing += 1
            availability_rows.append(
                {
                    "field": f,
                    "availability": klass,
                    "validated_seen": int(availability.get("validated_seen", 0) or 0),
                    "filled_rate_validated": float(availability.get("filled_rate_validated", 0) or 0),
                }
            )
        st.subheader("Field Availability Guidance")
        a1, a2, a3 = st.columns(3)
        a1.metric("Expected Missing", expected_missing, help="Expected fields should usually be found; hunt these first.")
        a2.metric("Sometimes Missing", sometimes_missing)
        a3.metric("Rare Missing", rare_missing, help="Rare fields are often not publicly disclosed.")
        with st.expander("Availability details for this product", expanded=False):
            st.dataframe(availability_rows, use_container_width=True, height=280)

    active_match = next((r for r in catalog["active"] if token(r["brand"]) == token(brand) and token(r["model"]) == token(model)), None)
    if active_match:
        st.subheader("Helper Target Snapshot")
        helper_known = {k: v for k, v in active_match["raw"].items() if k in set(order) and known(v)}
        total_schema_keys = max(len(order), 1)
        st.caption(
            f"activeFiltering populated schema keys: {len(helper_known)}/{total_schema_keys} "
            "(empty values in helper target will not fill specs)."
        )
        st.json(helper_known)

with tab3:
    filtered = events_for_product
    latest_run = latest_selected_run
    filtered_run = events_for_selected_run
    stage = pipeline_stage_status(filtered_run)
    counts = stage["counts"]
    is_running = bool(st.session_state.proc and st.session_state.proc.poll() is None)
    e1, e2, e3, e4 = st.columns(4)
    e1.metric("Events (Current Run)", len(filtered_run))
    e2.metric("Event Types", len(counts))
    e3.metric("Last Event", filtered_run[-1].get("event", "n/a") if filtered_run else "n/a")
    e4.metric("Event Log Bytes", EVENTS_PATH.stat().st_size if EVENTS_PATH.exists() else 0)
    st.caption(f"Run ID focus: `{latest_run or 'n/a'}`")
    st.progress(stage["progress_pct"] / 100.0, text=f"Pipeline Progress: {stage['progress_pct']:.0f}%")
    stage_rows = []
    active_assigned = False
    for item in stage["stages"]:
        if item["reached"]:
            state = "DONE"
        elif is_running and not active_assigned:
            state = "ACTIVE"
            active_assigned = True
        else:
            state = "PENDING"
        stage_rows.append(
            {
                "stage": item["label"],
                "status": state,
                "signals": ", ".join(item["events"]),
            }
        )
    st.dataframe(stage_rows, use_container_width=True, height=240)

    failure_total = sum(
        counts.get(name, 0)
        for name in (
            "llm_call_failed",
            "llm_extract_failed",
            "llm_discovery_planner_failed",
            "llm_summary_failed",
            "source_fetch_failed",
        )
    )
    f1, f2, f3, f4 = st.columns(4)
    f1.metric("Fetch Failures", counts.get("source_fetch_failed", 0))
    f2.metric("LLM Failures", counts.get("llm_call_failed", 0))
    f3.metric("Extraction Failures", counts.get("llm_extract_failed", 0))
    f4.metric("Total Failure Signals", failure_total)

    st.subheader("Event Legend")
    st.dataframe(event_rows_with_help(counts), use_container_width=True, height=260)

    tail = st.slider("Tail Events", min_value=20, max_value=1000, value=160, step=20)
    st.dataframe(filtered_run[-tail:], use_container_width=True, height=320)
    st.subheader("Process Output")
    st.caption(f"Process status: {'RUNNING' if is_running else 'IDLE'}")
    if st.session_state.last_cmd:
        st.caption(f"Last command: `{st.session_state.last_cmd}`")
    st.code("\n".join(st.session_state.proc_log[-260:]) or "No process output yet.", language="text")
    st.subheader("Queue Snapshot")
    qrows = list((catalog["queue"].get("products", {}) or {}).values())
    qrows.sort(key=lambda r: (str(r.get("status", "")), str(r.get("productId", ""))))
    st.dataframe(qrows[:400], use_container_width=True, height=300)

with tab4:
    monthly = load_billing_month() or {}
    totals = monthly.get("totals", {})
    b1, b2, b3, b4 = st.columns(4)
    b1.metric("Cost USD", f"{float(totals.get('cost_usd', 0)):.6f}")
    b2.metric("Calls", int(totals.get("calls", 0)))
    b3.metric("Prompt Tokens", int(totals.get("prompt_tokens", 0)))
    b4.metric("Completion Tokens", int(totals.get("completion_tokens", 0)))
    st.caption("By reason")
    st.json(monthly.get("by_reason", {}))
    st.caption("By model")
    st.json(monthly.get("by_model", {}))
    lroot = effective_learning_root() / category
    lfiles = []
    for name in ["field_lexicon.json", "constraints.json", "field_yield.json", "identity_grammar.json", "query_templates.json", "source_promotions.json", "stats.json"]:
        path = lroot / name
        lfiles.append({"file": name, "exists": path.exists(), "size_bytes": path.stat().st_size if path.exists() else 0, "updated_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat() if path.exists() else ""})
    st.subheader("Learning Artifacts")
    st.dataframe(lfiles, use_container_width=True, height=250)
    st.subheader("Component Library Counts")
    st.json({name: len(read_jsonl(COMPONENT_ROOT / f"{name}.jsonl", limit=10000)) for name in ("sensors", "switches", "encoders", "mcus")})
    st.subheader("Helper Files")
    st.caption("Active helper input: `helper_files/<category>/activeFiltering.json`")

with tab5:
    render_field_rules_studio(category, local)

st.caption("Use dropdowns from helper targets, run a product, and monitor events/fields/costs in real time.")
if st.session_state.auto_refresh and st.session_state.proc and st.session_state.proc.poll() is None:
    time.sleep(1.5)
    st.rerun()

